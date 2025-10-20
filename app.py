import sqlite3
from datetime import datetime, timedelta
from flask import json, render_template, redirect, url_for, flash
import pandas as pd
from pandas import isna
import logging
import zipfile
import io
from charset_normalizer import detect
import re
from flask import Flask, jsonify, request, session
import os
import shutil
import chardet
import csv
from dateutil.parser import parse
from flask import Flask, jsonify, request, session, render_template, redirect, url_for, flash, send_file
import zipfile
import tempfile
from PIL import Image
import io
from dateutil.parser import parse as parse_date
from dateutil.relativedelta import relativedelta
from collections import OrderedDict
from datetime import timedelta
import utils
import uuid
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import genscript
import zipfile
import threading
import time

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# SQLite database setup
DB_FILE = 'projects.db'
MAIL_DIR = 'mail'
PROJECT_DIR = 'project'
OLD_DIR = 'old'
FILEUPLOAD_DIR = 'fileupload'

DISPLAY_COLUMNS = [
    'ステータス', '案件名', 'PH','要件引継', '設計開始',
    '設計完了', '設計書送付', '開発開始', '開発完了', 'テスト開始日', 'テスト完了日',
    'FB完了予定日', 'SE納品', 'タスク', 'SE', 'SE(sub)', 'BSE', '案件番号', 'PJNo.', 
    '開発工数（h）', '設計工数（h）', 'ページ数', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト', '備考', '履歴'
]
DATE_COLUMNS_DB = [
    '要件引継', '設計開始', '設計完了', '設計書送付', '開発開始', '開発完了',
    'テスト開始日', 'テスト完了日', 'FB完了予定日', 'SE納品'
]
DATE_COLUMNS_DISPLAY = DATE_COLUMNS_DB.copy()
VALID_STATUSES = [
    '要件引継待ち', '設計中', 'SE送付済', '開発中', 'テスト中', 'FB対応中', 'SE納品済'
]

MAIL_DIR = 'mail'

generation_progress = {
    'current_sheet': 0,
    'total_sheets': 0,
    'sheet_name': '',
    'is_generating': False
}

# Excel comparison tracking
comparison_tasks = {}

def init_db():
    """Initialize SQLite database with projects, copied_templates, and daily_hours tables."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            SE TEXT,
            案件名 TEXT,
            PH TEXT,
            "開発工数（h）" REAL,
            "設計工数（h）" REAL,
            要件引継 TEXT,
            設計開始 TEXT,
            設計完了 TEXT,
            設計書送付 TEXT,
            開発開始 TEXT,
            開発完了 TEXT,
            SE納品 TEXT,
            BSE TEXT,
            案件番号 TEXT,
            "PJNo." TEXT,
            備考 TEXT,
            テスト開始日 TEXT,
            テスト完了日 TEXT,
            FB完了予定日 TEXT,
            ページ数 INTEGER,
            タスク TEXT,
            ステータス TEXT,
            不要 INTEGER DEFAULT 0,
            注文設計 INTEGER DEFAULT 0,
            注文テスト INTEGER DEFAULT 0,
            注文FB INTEGER DEFAULT 0,
            注文BrSE INTEGER DEFAULT 0,
            user_edited_status INTEGER DEFAULT 0
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS copied_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            filename TEXT,
            copied_at TEXT,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS daily_hours (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            date TEXT,
            task_type TEXT,
            hours REAL,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS todos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            date TEXT NOT NULL,
            priority TEXT DEFAULT 'medium',
            completed INTEGER DEFAULT 0,
            repeat_type TEXT DEFAULT 'none',
            repeat_interval INTEGER DEFAULT 1,
            repeat_unit TEXT DEFAULT 'days',
            end_date TEXT,
            parent_id INTEGER,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (parent_id) REFERENCES todos(id)
        )
    ''')
    
    cursor.execute("PRAGMA table_info(todos)")
    todo_columns = [col[1] for col in cursor.fetchall()]
    if 'project_id' not in todo_columns:
        cursor.execute('ALTER TABLE todos ADD COLUMN project_id INTEGER')
        
    cursor.execute("PRAGMA table_info(projects)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'ステータス' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN ステータス TEXT')
        cursor.execute("UPDATE projects SET ステータス = '要件引継待ち' WHERE ステータス IS NULL")
    if '不要' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN 不要 INTEGER DEFAULT 0')
    if '注文設計' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN 注文設計 INTEGER DEFAULT 0')
    if '注文テスト' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN 注文テスト INTEGER DEFAULT 0')
    if '注文FB' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN 注文FB INTEGER DEFAULT 0')
    if '注文BrSE' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN 注文BrSE INTEGER DEFAULT 0')
    if 'user_edited_status' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN user_edited_status INTEGER DEFAULT 0')
    if 'SE(sub)' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN "SE(sub)" TEXT')
    if '並行テスト' not in columns:
        cursor.execute('ALTER TABLE projects ADD COLUMN 並行テスト INTEGER DEFAULT 0')

    cursor.execute('''
            CREATE TABLE IF NOT EXISTS schedule_done_status (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER,
                date_column TEXT,
                done INTEGER DEFAULT 0,
                FOREIGN KEY (project_id) REFERENCES projects(id)
            )
        ''')
    
    # Thêm bảng editor_document mới
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS editor_document (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
     # Add memo table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS memo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
     # Add memo_files table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS memo_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            memo_id INTEGER,
            filename TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            file_type TEXT NOT NULL,
            file_size INTEGER,
            uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (memo_id) REFERENCES memo(id)
        )
    ''')
    
    # Add project_history table for tracking all changes to projects
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            action_type TEXT NOT NULL,
            action_details TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')


    conn.commit()
    conn.close()

def read_users():
    """Read users from users.txt, create default admin if not exists."""
    users = {}
    try:
        with open('users.txt', 'r', encoding='utf-8') as f:
            for line in f:
                username, password = line.strip().split(':')
                users[username] = password
    except FileNotFoundError:
        with open('users.txt', 'w', encoding='utf-8') as f:
            f.write('admin:admin123\n')
        users['admin'] = 'admin123'
    return users

def add_project_history(project_id, action_type, action_details=''):
    """Add or update history entry for a project (keeps only the latest entry per project).
    
    Args:
        project_id: ID of the project
        action_type: Type of action (e.g., 'created', 'updated', 'mail_sent', 'copied')
        action_details: Details about the action (e.g., field names, template name)
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Check if history already exists for this project
        cursor.execute('SELECT id FROM project_history WHERE project_id = ?', (project_id,))
        existing = cursor.fetchone()
        
        if existing:
            # Update existing history entry
            cursor.execute('''
                UPDATE project_history
                SET action_type = ?, action_details = ?, created_at = datetime('now', 'localtime')
                WHERE project_id = ?
            ''', (action_type, action_details, project_id))
            logging.info(f"Updated history for project {project_id}: {action_type} - {action_details}")
        else:
            # Insert new history entry
            cursor.execute('''
                INSERT INTO project_history (project_id, action_type, action_details, created_at)
                VALUES (?, ?, ?, datetime('now', 'localtime'))
            ''', (project_id, action_type, action_details))
            logging.info(f"Added history for project {project_id}: {action_type} - {action_details}")
        
        conn.commit()
        conn.close()
    except Exception as e:
        logging.error(f"Error adding/updating project history: {e}")

def project_exists(cursor, project):
    """Check if a project already exists based on 案件名, PH, PJNo., 案件番号.
    Returns tuple (existing_project, match_level) if found, (None, 0) otherwise.
    Match levels:
    - 4: All 4 keys match (案件名, PH, PJNo., 案件番号) - Full match, Update allowed
    - 3: 3 keys match (案件名, PH, PJNo.) - Good match, Update allowed
    - 2: 2 keys match (案件名, PH) - Partial match, Update NOT allowed
    - 1: Only 1 key matches - Weak match, Update NOT allowed
    Excludes projects marked as 不要 (treated as deleted/ignored)."""
    
    # Define key groups in priority order
    all_keys = ['案件名', 'PH', 'PJNo.', '案件番号']
    key_groups = [
        (['案件名', 'PH', 'PJNo.', '案件番号'], 4),  # All 4 keys
        (['案件名', 'PH', 'PJNo.'], 3),              # 3 keys
        (['案件名', 'PH'], 2),                       # 2 keys
    ]
    
    # Exclude projects marked as 不要 = ON (1 / "1" / "true")
    exclude_unnecessary = " AND (不要 IS NULL OR 不要 = 0 OR 不要 = '0' OR 不要 = 'false' OR 不要 = 'False')"
    
    # Try matching with each key group in priority order
    for keys, match_level in key_groups:
        conditions = []
        values = []
        
        for key in keys:
            value = project.get(key, '')
            if value is None or (isinstance(value, str) and value.strip() == '') or isna(value):
                continue
            conditions.append(f'"{key}" = ?')
            values.append(str(value))
        
        # Skip this key group if not all keys are present
        if len(conditions) != len(keys):
            continue
        
        query = f'''
            SELECT * FROM projects
            WHERE {' AND '.join(conditions)} {exclude_unnecessary}
        '''
        
        logging.debug(f"Trying match_level {match_level}: {query} with values: {values}")
        cursor.execute(query, values)
        row = cursor.fetchone()
        
        if row:
            columns = [description[0] for description in cursor.description]
            existing_project = dict(zip(columns, row))
            logging.debug(f"Found existing project with id: {existing_project.get('id')} at match_level {match_level}")
            return existing_project, match_level
    
    # Check for single key matches (match_level = 1)
    for key in all_keys:
        value = project.get(key, '')
        if value is None or (isinstance(value, str) and value.strip() == '') or isna(value):
            continue
        
        query = f'''
            SELECT * FROM projects
            WHERE "{key}" = ? {exclude_unnecessary}
        '''
        
        logging.debug(f"Trying single key match for {key}: {query} with value: {value}")
        cursor.execute(query, [str(value)])
        row = cursor.fetchone()
        
        if row:
            columns = [description[0] for description in cursor.description]
            existing_project = dict(zip(columns, row))
            logging.debug(f"Found existing project with id: {existing_project.get('id')} at match_level 1 (key: {key})")
            return existing_project, 1
    
    logging.debug("No matching project found")
    return None, 0

def compare_projects(existing_project, new_project):
    """Compare two projects and return differences.
    Returns dict with field names as keys and {'old': value, 'new': value} as values."""
    differences = {}
    
    # List of fields to compare (exclude id, internal fields, and calculated fields)
    # Exclude ステータス because it's calculated automatically, not from Excel
    excluded_fields = ['id', 'user_edited_status', 'ステータス']
    fields_to_compare = [col for col in DISPLAY_COLUMNS if col in new_project and col not in excluded_fields]
    
    for field in fields_to_compare:
        old_value = existing_project.get(field, '')
        new_value = new_project.get(field, '')
        
        # Normalize values for comparison - handle NaN properly
        if old_value is None or (isinstance(old_value, float) and pd.isna(old_value)):
            old_value = ''
        if new_value is None or (isinstance(new_value, float) and pd.isna(new_value)):
            new_value = ''
            
        # Convert to string for comparison, handle 'nan' string
        old_str = str(old_value).strip()
        new_str = str(new_value).strip()
        
        # Clean up 'nan' strings that come from float conversion
        if old_str.lower() == 'nan':
            old_str = ''
        if new_str.lower() == 'nan':
            new_str = ''
        
        # Check if values are different
        if old_str != new_str:
            differences[field] = {
                'old': old_str,
                'new': new_str
            }
    
    return differences


def parse_date_from_db(date_str):
    """Parse date string from database (YYYY-MM-DD or YYYY/MM/DD) to datetime object."""
    if isna(date_str) or date_str is None or date_str == '':
        logging.debug(f"Date string is empty or None: {date_str}")
        return None
    try:
        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
        logging.debug(f"Successfully parsed date: {date_str} -> {parsed_date}")
        return parsed_date
    except ValueError:
        try:
            parsed_date = datetime.strptime(date_str, '%Y/%m/%d')
            logging.debug(f"Successfully parsed date: {date_str} -> {parsed_date}")
            return parsed_date
        except (ValueError, TypeError) as e:
            #logging.error(f"Failed to parse date: {date_str}, error: {e}")
            return None

def parse_date_for_comparison(date_str):
    """Parse date string for comparison, supports YYYY/MM/DD(曜日) and YYYY-MM-DD."""
    if isna(date_str) or date_str is None or date_str == '':
        logging.debug(f"Date string for comparison is empty or None: {date_str}")
        return None
    try:
        if isinstance(date_str, datetime):
            return date_str
        if '(' in date_str:
            date_str = date_str.split('(')[0]
        parsed_date = datetime.strptime(date_str, '%Y/%m/%d')
        logging.debug(f"Successfully parsed date for comparison: {date_str} -> {parsed_date}")
        return parsed_date
    except ValueError:
        try:
            parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
            logging.debug(f"Successfully parsed date for comparison: {date_str} -> {parsed_date}")
            return parsed_date
        except (ValueError, TypeError) as e:
            #logging.error(f"Failed to parse date for comparison: {date_str}, error: {e}")
            return None

def format_date_jp(date):
    """Format datetime object to YYYY/MM/DD(曜日)."""
    if date is None:
        return ''
    weekdays = ['月', '火', '水', '木', '金', '土', '日']
    weekday = weekdays[date.weekday()]
    return date.strftime('%Y/%m/%d') + f'({weekday})'

def convert_nat_to_none(project_dict):
    """Convert NaT/NaN/None values to empty strings and handle specific data types."""
    for key, value in project_dict.items():
        if isna(value) or value is None:
            project_dict[key] = '' if key not in ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト',
                                                  'user_edited_status'] else 0
        elif key == 'PJNo.':
            if isinstance(value, (float, int)):
                project_dict[key] = str(int(value))
            else:
                project_dict[key] = str(value)
        elif isinstance(value, (float, int)) and key not in ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE',
                                                             'user_edited_status']:
            project_dict[key] = str(value)
        elif key in ['注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト']:
            project_dict[key] = '○' if value == 1 else ''
        if key == 'fb_late':
            project_dict[key] = bool(value)
    return project_dict

def calculate_status(project, current_date=None):
    """Calculate project status based on milestone dates and current date, unless user_edited_status is 1."""
    if project.get('user_edited_status', 0) == 1:
        return project.get('ステータス', '要件引継待ち')

    if current_date is None:
        current_date = datetime.now()

    date_fields = [
        ('要件引継', '要件引継待ち'),
        ('設計完了', '設計中'),
        ('設計書送付', 'SE送付済'),
        ('開発完了', '開発中'),
        ('テスト完了日', 'テスト中'),
        ('SE納品', 'FB対応中'),
    ]

    current_date = current_date.date()

    se_delivery_date = parse_date_from_db(project.get('SE納品', ''))
    if se_delivery_date and se_delivery_date.date() < current_date:
        return 'SE納品済'

    for date_field, status in date_fields:
        date_value = parse_date_from_db(project.get(date_field, ''))
        if date_value and date_value.date() >= current_date:
            return status

    return '要件引継待ち'

def read_pages_ranges():
    """Read page ranges and corresponding days from pages.txt."""
    ranges = []
    try:
        with open('pages.txt', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    range_part, days = line.split(':')
                    days = int(days.strip())
                    from_page, to_page = map(int, range_part.split('-'))
                    ranges.append((from_page, to_page, days))
                except ValueError:
                    continue
    except FileNotFoundError:
        pass
    return ranges

def read_working_days(file_path='config.txt'):
    """Read working days from config.txt, default to 9 if not found."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('fix FB days =') or line.startswith('const workingDays ='):
                    try:
                        value = int(line.split('=')[1].strip())
                        return value
                    except (IndexError, ValueError):
                        #logging.error(f"Invalid workingDays format in {file_path}: {line}")
                        return 9
        logging.warning(f"workingDays not found in {file_path}")
        return 9
    except FileNotFoundError:
        logging.warning(f"File {file_path} not found")
        return 9
    except Exception as e:
        #logging.error(f"Error reading {file_path}: {e}")
        return 9

def add_working_days(start_date, working_days):
    """Add working days to a start date, skipping weekends."""
    if not start_date or working_days <= 0:
        return ''
    current_date = start_date
    days_added = 0
    while days_added < working_days:
        current_date += timedelta(days=1)
        if current_date.weekday() < 5:
            days_added += 1
    return current_date.strftime('%Y-%m-%d')

def calculate_test_completion_date(page_count, test_start_date):
    logging.debug(f"Calculating test completion date: page_count={page_count}, test_start_date={test_start_date}")
    if not page_count or not test_start_date:
        #logging.error("Invalid inputs: page_count or test_start_date is empty")
        return ''
    try:
        page_count = int(page_count)
        test_start_date = datetime.strptime(test_start_date, '%Y-%m-%d')
    except (ValueError, TypeError) as e:
        #logging.error(f"Input parsing error: {str(e)}")
        return ''

    ranges = read_pages_ranges()
    logging.debug(f"Page ranges from pages.txt: {ranges}")
    for from_page, to_page, days in ranges:
        if from_page <= page_count <= to_page:
            result = add_working_days(test_start_date, days)
            logging.debug(f"Test completion date calculated: {result}")
            return result
    logging.warning(f"No matching range found for page_count={page_count}")
    return ''


def calculate_fb_completion_date(test_completion_date):
    logging.debug(f"Calculating FB completion date: test_completion_date={test_completion_date}")
    if not test_completion_date:
        logging.error("Test completion date is empty")
        return ''
    try:
        if isinstance(test_completion_date, str):
            test_completion = datetime.strptime(test_completion_date, '%Y-%m-%d')
        elif isinstance(test_completion_date, datetime):
            test_completion = test_completion_date
        else:
            raise ValueError("Invalid test_completion_date type")

        working_days = read_working_days()
        result = add_working_days(test_completion, working_days)
        logging.debug(f"FB completion date calculated: {result}")
        return result
    except (ValueError, TypeError) as e:
        logging.error(f"Error calculating FB completion date: {str(e)}")
        return ''

def import_excel_to_sqlite(file_path):
    """Import projects from Excel file to SQLite database."""
    if not os.path.exists(file_path):
        return False

    try:
        df = pd.read_excel(file_path, engine='openpyxl', dtype={'PJNo.': str})
        available_columns = [col for col in DISPLAY_COLUMNS if col in df.columns]
        df = df[available_columns].copy()

        if 'SE(sub)' not in df.columns:
            df['SE(sub)'] = ''

        current_date = datetime.now()

        # Process date columns
        for col in DATE_COLUMNS_DB:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')
            else:
                df[col] = ''


        # Set test start date based on development completion
        if '開発完了' in df.columns:
            def calculate_test_start_date(dev_complete):
                if pd.notna(dev_complete) and dev_complete != '':
                    try:
                        dev_complete_date = pd.to_datetime(dev_complete)
                        test_start_date = dev_complete_date + timedelta(days=1)
                        # Kiểm tra nếu là cuối tuần
                        if test_start_date.weekday() == 5:  # Thứ Bảy
                            test_start_date += timedelta(days=2)
                        elif test_start_date.weekday() == 6:  # Chủ Nhật
                            test_start_date += timedelta(days=1)
                        return test_start_date.strftime('%Y-%m-%d')
                    except (ValueError, TypeError):
                        return ''
                return ''

            df['テスト開始日'] = df['開発完了'].apply(calculate_test_start_date)

        # Handle PJNo. formatting
        if 'PJNo.' in df.columns:
            df['PJNo.'] = df['PJNo.'].apply(
                lambda x: str(int(float(x))) if pd.notna(x) and x != '' and isinstance(x, (int, float)) else str(
                    x) if pd.notna(x) else ''
            )

        # Handle PH formatting
        if 'PH' in df.columns:
            df['PH'] = df['PH'].apply(
                lambda x: str(int(float(x))) if pd.notna(x) and x != '' and isinstance(x, (int, float)) else str(
                    x) if pd.notna(x) else ''
            )

        # Process 案件名: Remove "株式会社"
        if '案件名' in df.columns:
            df['案件名'] = df['案件名'].apply(
                lambda x: str(x).replace('株式会社', '') if pd.notna(x) and isinstance(x, str) else str(x)
            )

        # Process SE: Keep only the part before the first space (1-byte or 2-byte)
        if 'SE' in df.columns:
            df['SE'] = df['SE'].apply(
                lambda x: str(x).split()[0] if pd.notna(x) and isinstance(x, str) and str(x).strip() else str(x)
            )

        if 'SE(sub)' in df.columns:
            df['SE(sub)'] = df['SE(sub)'].apply(
                lambda x: str(x).split()[0] if pd.notna(x) and isinstance(x, str) and str(x).strip() else str(x)
            )


        # Initialize missing columns
        for col in ['ページ数', 'タスク', 'ステータス', '不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト',
                    'user_edited_status']:
            if col not in df.columns:
                df[col] = '' if col != 'ステータス' else '要件引継待ち'
                if col in ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト', 'user_edited_status']:
                    df[col] = 0

        # Calculate status for each project
        for index, row in df.iterrows():
            project = row.to_dict()
            df.at[index, 'ステータス'] = calculate_status(project, current_date)
            

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        duplicated_projects = []
        imported_count = 0
        total_projects = len(df)
        logging.debug(f"Total projects to import: {total_projects}")
        
        for _, row in df.iterrows():
            try:
                project = row.to_dict()
                existing_project, match_level = project_exists(cursor, project)
                
                if not existing_project:
                    # New project - insert directly
                    cursor.execute('''
                                INSERT INTO projects (
                                    SE, "SE(sub)", 案件名, PH, "開発工数（h）", "設計工数（h）", 要件引継, 設計開始,
                                    設計完了, 設計書送付, 開発開始, 開発完了, SE納品, BSE, 案件番号, "PJNo.",
                                    備考, テスト開始日, テスト完了日, FB完了予定日, ページ数, タスク, ステータス,
                                    不要, 注文設計, 注文テスト, 注文FB, 注文BrSE, user_edited_status
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                        project.get('SE', ''),
                        project.get('SE(sub)', ''),
                        project.get('案件名', ''),
                        project.get('PH', ''),
                        project.get('開発工数（h）', None),
                        project.get('設計工数（h）', None),
                        project.get('要件引継', ''),
                        project.get('設計開始', ''),
                        project.get('設計完了', ''),
                        project.get('設計書送付', ''),
                        project.get('開発開始', ''),
                        project.get('開発完了', ''),
                        project.get('SE納品', ''),
                        project.get('BSE', ''),
                        project.get('案件番号', ''),
                        project.get('PJNo.', ''),
                        project.get('備考', ''),
                        project.get('テスト開始日', ''),
                        project.get('テスト完了日', ''),
                        project.get('FB完了予定日', ''),
                        project.get('ページ数', None),
                        project.get('タスク', ''),
                        project.get('ステータス', '要件引継待ち'),
                        project.get('不要', 0),
                        project.get('注文設計', 0),
                        project.get('注文テスト', 0),
                        project.get('注文FB', 0),
                        project.get('注文BrSE', 0),
                        project.get('user_edited_status', 0)
                    ))
                    new_project_id = cursor.lastrowid
                    # Add history entry for new project
                    cursor.execute('''
                        INSERT INTO project_history (project_id, action_type, action_details, created_at)
                        VALUES (?, 'created', 'Excelから新規作成', datetime('now', 'localtime'))
                    ''', (new_project_id,))
                    imported_count += 1
                else:
                    # Duplicate found - check if there are actual differences
                    differences = compare_projects(existing_project, project)
                    
                    # Only add to duplicated_projects if there are differences
                    if differences:
                        # Clean project data - convert NaN to None/empty string for JSON serialization
                        clean_project = {}
                        for key, value in project.items():
                            if pd.isna(value):
                                # Use None for numeric fields, empty string for text fields
                                if key in ['開発工数（h）', '設計工数（h）', 'ページ数']:
                                    clean_project[key] = None
                                elif key in ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト', 'user_edited_status']:
                                    clean_project[key] = 0
                                else:
                                    clean_project[key] = ''
                            else:
                                clean_project[key] = value
                        
                        duplicated_projects.append({
                            'project_name': clean_project.get('案件名', ''),
                            'ph': clean_project.get('PH', ''),
                            'pjno': clean_project.get('PJNo.', ''),
                            'existing_id': existing_project['id'],
                            'differences': differences,
                            'new_data': clean_project,
                            'match_level': match_level  # Add match_level to frontend
                        })
                    else:
                        # Duplicate but no differences - skip silently (already in DB)
                        logging.debug(f"Skipping duplicate project with no differences: {project.get('案件名', '')} (PH: {project.get('PH', '')}, PJNo: {project.get('PJNo.', '')})")
            except Exception as e:
                logging.error(f"Lỗi ở dòng import project: {e}")
                

        logging.info(f'Found {len(duplicated_projects)} duplicated projects')
        conn.commit()
        conn.close()
        logging.info(f"Imported {imported_count} new projects from {file_path}")
        return True, duplicated_projects, total_projects, imported_count
    except Exception as e:
        #logging.error(f"Failed to import Excel file {file_path}: {e}")
        return False, [], 0, 0

def read_projects():
    """Read all projects from SQLite database with total hours worked and latest history."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM projects')
    rows = cursor.fetchall()
    columns = [description[0] for description in cursor.description]

    projects = []
    for row in rows:
        project = dict(zip(columns, row))
        
        # Get total hours by task type
        cursor.execute('''
            SELECT task_type, SUM(hours) as total_hours
            FROM daily_hours
            WHERE project_id = ?
            GROUP BY task_type
        ''', (project['id'],))
        hours = {row[0]: row[1] for row in cursor.fetchall()}
        project['設計実績'] = hours.get('設計', 0)
        project['テスト実績'] = hours.get('テスト', 0)
        project['FB実績'] = hours.get('FB', 0)
        project['BrSE実績'] = hours.get('BrSE', 0)
        
        # Get latest history entry for this project
        cursor.execute('''
            SELECT action_type, action_details, created_at
            FROM project_history
            WHERE project_id = ?
            ORDER BY created_at DESC
            LIMIT 1
        ''', (project['id'],))
        history_row = cursor.fetchone()
        
        if history_row:
            action_type, action_details, created_at = history_row
            # Format history display based on action type
            if action_type == 'created':
                project['履歴'] = '新規作成'
            elif action_type == 'updated':
                project['履歴'] = f'{action_details}を更新' if action_details else '更新'
            elif action_type == 'excel_updated':
                project['履歴'] = f'{action_details}を更新 (Excel)'
            elif action_type == 'mail_sent':
                project['履歴'] = f'メール送信: {action_details}'
            elif action_type == 'copied':
                project['履歴'] = 'プロジェクトコピー'
            else:
                project['履歴'] = action_details or action_type
        else:
            project['履歴'] = ''
        
        projects.append(project)

    conn.close()
    df = pd.DataFrame(projects)
    return df

def get_mail_templates():
    """Get list of mail templates from mail directory."""
    if not os.path.exists(MAIL_DIR):
        os.makedirs(MAIL_DIR)
    templates = [f for f in os.listdir(MAIL_DIR) if f.endswith('.txt')]
    templates = [(f, f[:-4]) for f in sorted(templates)]
    #logging.debug(f"Mail templates: {templates}")
    return templates

def get_week_dates(week_start):
    """Tạo danh sách ngày trong tuần bắt đầu từ Monday."""
    dates = []
    current_date = datetime.strptime(week_start, '%Y-%m-%d')
    for i in range(7):
        date = current_date + timedelta(days=i)
        weekdays = ['月', '火', '水', '木', '金', '土', '日']
        display = date.strftime('%Y/%m/%d') + f'({weekdays[date.weekday()]})'
        dates.append({'date': date.strftime('%Y-%m-%d'), 'display': display})
    return dates

def update_project(project_id, updates):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    current_date = datetime.now()

    if 'SE(sub)' in updates:
        updates['SE(sub)'] = updates['SE(sub)'].split()[0] if updates['SE(sub)'] and isinstance(updates['SE(sub)'],
                                                                                                str) else updates[
            'SE(sub)']

    if 'ページ数' in updates:
        if updates['ページ数'] == '':
            updates['ページ数'] = None
        else:
            try:
                page_count = int(updates['ページ数'])
                if page_count <= 0:
                    raise ValueError
                updates['ページ数'] = page_count
            except ValueError:
                updates['ページ数'] = None

    if 'タスク' in updates:
        valid_tasks = ['設計', 'Brse', 'テスト', 'FB']
        tasks = updates['タスク'].split(',') if updates['タスク'] else []
        tasks = [task.strip() for task in tasks if task.strip() in valid_tasks]
        updates['タスク'] = ','.join(tasks) if tasks else ''
    else:
        updates['タスク'] = ''

    for field in ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト']:
        if field in updates:
            updates[field] = 1 if updates[field] == 'on' else 0

    # If 不要 is set to ON (1), delete all todos for this project
    if updates.get('不要') == 1:
        cursor.execute('DELETE FROM todos WHERE project_id = ?', (project_id,))
        logging.info(f"Deleted all todos for project {project_id} because 不要 was set to ON")

    # Handle ステータス and user_edited_status
    if 'ステータス' in updates and updates['ステータス'] in VALID_STATUSES:
        updates['user_edited_status'] = 1
    else:
        updates['user_edited_status'] = 0

    if '開発完了' in updates and updates['開発完了']:
        try:
            dev_complete_date = datetime.strptime(updates['開発完了'], '%Y-%m-%d')
            test_start_date = dev_complete_date + timedelta(days=1)
            # Kiểm tra nếu là cuối tuần
            if test_start_date.weekday() == 5:  # Thứ Bảy
                test_start_date += timedelta(days=2)
            elif test_start_date.weekday() == 6:  # Chủ Nhật
                test_start_date += timedelta(days=1)
            updates['テスト開始日'] = test_start_date.strftime('%Y-%m-%d')
        except ValueError:
            updates['テスト開始日'] = ''
    elif '開発完了' in updates and not updates['開発完了']:
        updates['テスト開始日'] = ''

    # Calculate テスト完了日 and FB完了予定日
    page_count = updates.get('ページ数')
    test_start_date = updates.get('テスト開始日')
    #logging.debug(f"Calculating test dates: page_count={page_count}, test_start_date={test_start_date}")

    # Use values from form if available and valid
    if 'テスト完了日' in updates and updates['テスト完了日']:
        try:
            datetime.strptime(updates['テスト完了日'], '%Y-%m-%d')
            #logging.debug(f"Using テスト完了日 from form: {updates['テスト完了日']}")
        except ValueError:
            updates['テスト完了日'] = ''
    else:
        updates['テスト完了日'] = calculate_test_completion_date(page_count, test_start_date) if page_count and test_start_date else ''

    if 'FB完了予定日' in updates and updates['FB完了予定日']:
        try:
            datetime.strptime(updates['FB完了予定日'], '%Y-%m-%d')
            #logging.debug(f"Using FB完了予定日 from form: {updates['FB完了予定日']}")
        except ValueError:
            updates['FB完了予定日'] = ''
    else:
        updates['FB完了予定日'] = calculate_fb_completion_date(updates['テスト完了日']) if updates['テスト完了日'] else ''

    # Only calculate status if not user-edited
    if updates.get('user_edited_status', 0) == 0:
        updates['ステータス'] = calculate_status(updates, current_date)

    # Get old values from database to compare
    cursor.execute('SELECT * FROM projects WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    if row:
        columns = [description[0] for description in cursor.description]
        old_project = dict(zip(columns, row))
    else:
        old_project = {}
    
    # Get list of actually changed fields (excluding internal fields)
    excluded_fields = ['user_edited_status', 'ステータス', 'id']
    numeric_fields = ['開発工数（h）', '設計工数（h）', 'ページ数']
    checkbox_fields = ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト']
    changed_fields = []
    
    for key, new_value in updates.items():
        if key in excluded_fields:
            continue
        
        old_value = old_project.get(key, '')
        
        # Special handling for numeric fields
        if key in numeric_fields:
            # Convert both to float for comparison, treat None/empty as 0
            try:
                old_num = float(old_value) if old_value not in [None, '', 'None'] else 0
            except (ValueError, TypeError):
                old_num = 0
            
            try:
                new_num = float(new_value) if new_value not in [None, '', 'None'] else 0
            except (ValueError, TypeError):
                new_num = 0
            
            if old_num != new_num:
                changed_fields.append(key)
        
        # Special handling for checkbox fields
        elif key in checkbox_fields:
            old_bool = int(old_value) if old_value not in [None, '', 'None'] else 0
            new_bool = int(new_value) if new_value not in [None, '', 'None'] else 0
            
            if old_bool != new_bool:
                changed_fields.append(key)
        
        # Regular string comparison for other fields
        else:
            # Normalize values for comparison
            if old_value is None or (isinstance(old_value, float) and pd.isna(old_value)):
                old_value = ''
            if new_value is None or (isinstance(new_value, float) and pd.isna(new_value)):
                new_value = ''
            
            # Convert to string for comparison
            old_str = str(old_value).strip() if old_value != '' else ''
            new_str = str(new_value).strip() if new_value != '' else ''
            
            # Only add to changed_fields if values are actually different
            if old_str != new_str:
                changed_fields.append(key)
    
    set_clause_parts = []
    values = []
    for key, value in updates.items():
        set_clause_parts.append(f'"{key}" = ?')
        values.append(value)
    set_clause = ', '.join(set_clause_parts)
    values.append(project_id)

    sql = f'''
        UPDATE projects
        SET {set_clause}
        WHERE id = ?
    '''
    #logging.debug(f"Executing SQL: {sql} with values: {values}")
    print(sql)
    cursor.execute(sql, values)
    
    # Add or update history entry for updated fields
    if changed_fields:
        # Limit to first 5 fields for display
        display_fields = changed_fields[:5]
        if len(changed_fields) > 5:
            action_details = ', '.join(display_fields) + f' 他{len(changed_fields) - 5}項目'
        else:
            action_details = ', '.join(display_fields)
        
        # Check if history exists and update or insert
        cursor.execute('SELECT id FROM project_history WHERE project_id = ?', (project_id,))
        existing = cursor.fetchone()
        
        if existing:
            cursor.execute('''
                UPDATE project_history
                SET action_type = ?, action_details = ?, created_at = datetime('now', 'localtime')
                WHERE project_id = ?
            ''', ('updated', action_details, project_id))
        else:
            cursor.execute('''
                INSERT INTO project_history (project_id, action_type, action_details, created_at)
                VALUES (?, 'updated', ?, datetime('now', 'localtime'))
            ''', (project_id, action_details))
    
    conn.commit()
    conn.close()

@app.route('/')
def index():
    """Redirect to dashboard if logged in, else to login."""
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login."""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        users = read_users()
        if username in users and users[username] == password:
            session['username'] = username
            flash('ログイン成功', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('ユーザー名またはパスワードが正しくありません', 'danger')
    return render_template('login.html')

# Thêm từ điển ánh xạ trạng thái với mức độ ưu tiên
STATUS_PRIORITY = {
    '要件引継待ち': 6,
    '設計中': 1,
    'SE送付済': 2,
    '開発中': 3,
    'テスト中': 4,
    'FB対応中': 5,
    'SE納品済': 7
}

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    """Render dashboard with project data and handle project updates."""
    if 'username' not in session:
        return redirect(url_for('login'))

    init_db()

    if not os.path.exists(PROJECT_DIR):
        os.makedirs(PROJECT_DIR)
    if not os.path.exists(OLD_DIR):
        os.makedirs(OLD_DIR)
    if not os.path.exists('temp_compare'):
        os.makedirs('temp_compare')

    show_all = request.form.get('show_all') == 'on'
    show_unnecessary = request.form.get('show_unnecessary') == 'on'

    df = read_projects()
    current_date = datetime.now()
    filtered_projects = []

    for _, row in df.iterrows():
        # Apply filtering: include project if (不要 = 0 OR show_unnecessary) AND (SE納品 not past OR show_all)
        if not show_unnecessary and row.get('不要', 0) == 1:
            continue

        # se_delivery_date = parse_date_from_db(row['SE納品'])
        # if not show_all and se_delivery_date and se_delivery_date.date() < current_date.date():
        #     continue

        # if not show_all:
        #     continue

        project = row.to_dict()
        project['ステータス'] = calculate_status(project, current_date)

        closest_date = None
        min_diff = float('inf')
        for col in DATE_COLUMNS_DISPLAY:
            date_str_db = row[col]
            date_obj = parse_date_from_db(date_str_db)
            project[f'{col}_past'] = False
            if date_obj is not None:
                diff = (date_obj.date() - current_date.date()).days
                if diff >= 0 and diff < min_diff:
                    min_diff = diff
                    closest_date = col
                if date_obj.date() < current_date.date():
                    project[f'{col}_past'] = True
            project[col] = format_date_jp(date_obj)
        project['highlight_column'] = closest_date if closest_date else None
        
        fb_completion_date = parse_date_from_db(row['FB完了予定日'])
        project['fb_late'] = False

        se_delivery_date = parse_date_from_db(row['SE納品'])
        if fb_completion_date and se_delivery_date:
            project['fb_late'] = fb_completion_date.date() > se_delivery_date.date()

        project = convert_nat_to_none(project)
        filtered_projects.append(project)

    # Sắp xếp mặc định theo yêu cầu
    def safe_get(project, key):
        if key == 'ステータス':
            # Trả về mức độ ưu tiên của trạng thái
            return STATUS_PRIORITY.get(project.get(key, '要件引継待ち'), 8)
        elif key in ['設計開始', '設計完了']:
            # Xử lý ngày, trả về datetime.max nếu trống để xếp cuối khi tăng dần
            date_obj = parse_date_for_comparison(project.get(key, ''))
            return date_obj if date_obj else datetime.max
        elif key == '設計工数（h）':
            # Xử lý số thực, trả về 0 nếu trống để xếp cuối khi giảm dần
            try:
                return float(project.get(key, 0)) if project.get(key, '') != '' else 0
            except ValueError:
                return 0
        return project.get(key, '')

    filtered_projects.sort(
        key=lambda x: (
            safe_get(x, 'ステータス'),           # 1. ステータス (theo thứ tự ưu tiên)
            safe_get(x, '設計開始'),            # 2. 設計開始 (tăng dần)
            safe_get(x, '設計完了'),            # 3. 設計完了 (tăng dần)
            -safe_get(x, '設計工数（h）')       # 4. 設計工数（h） (giảm dần, dùng dấu - để đảo ngược)
        )
    )

    if request.method == 'POST' and 'index' in request.form:
        try:
            project_id = int(request.form['index'])
        except ValueError:
            flash('エラー: 無効なプロジェクトIDです', 'danger')
            return redirect(url_for('dashboard'))

        updates = {}
        for col in df.columns:
            if col in request.form and col != 'id':
                if col == 'タスク':
                    updates[col] = ','.join(request.form.getlist(col))
                else:
                    updates[col] = request.form[col]
            # Đảm bảo các trường checkbox luôn có mặt trong updates
        for field in ['不要', '注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト']:
            if field not in updates:
                updates[field] = 0
            
        update_project(project_id, updates)
        project_name = updates.get('案件名', '')
        ph = updates.get('PH', '')
        new_start = updates.get('設計開始', '')
        new_end = updates.get('設計完了', '')
        
        print("project keys:", project.keys())
        print("project['PJNo.']:", project.get('PJNo.'), "project['案件番号']:", project.get('案件番号'))
                            
        pjno = str(project.get('PJNo.') or project.get('案件番号') or '').strip()
        if project_name and new_start and new_end:
            sync_design_todo(project_id, project_name, ph, new_start, new_end, pjno)
        flash('プロジェクトが正常に更新されました', 'success')
        return redirect(url_for('dashboard'))

    mail_templates = get_mail_templates()
    ranges = read_pages_ranges()
    working_days = read_working_days()

    return render_template('dashboard.html',
                           projects=filtered_projects,
                           display_columns=DISPLAY_COLUMNS,
                           date_columns=DATE_COLUMNS_DISPLAY,
                           show_all=show_all,
                           show_unnecessary=show_unnecessary,
                           mail_templates=mail_templates,
                           ranges=ranges,
                           valid_statuses=VALID_STATUSES,
                           working_days=working_days)

@app.route('/upload', methods=['POST'])
def upload():
    """Handle Excel file upload."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'error': '許可されていないファイル形式です'}), 400

    try:
        if not os.path.exists(PROJECT_DIR):
            os.makedirs(PROJECT_DIR)
        if not os.path.exists(OLD_DIR):
            os.makedirs(OLD_DIR)

        file_path = os.path.join(PROJECT_DIR, file.filename)
        file.save(file_path)

        for existing_file in os.listdir(PROJECT_DIR):
            if existing_file != file.filename:
                existing_file_path = os.path.join(PROJECT_DIR, existing_file)
                if os.path.isfile(existing_file_path):
                    os.remove(existing_file_path)

        result, duplicated_projects, total_projects, imported_count = import_excel_to_sqlite(file_path)
        os.remove(file_path)
        
        if result:
            if duplicated_projects:
                # Have duplicates with differences - return them for user to decide
                return jsonify({
                    'success': True,
                    'has_duplicates': True,
                    'imported_count': imported_count,
                    'duplicate_count': len(duplicated_projects),
                    'duplicates': duplicated_projects
                })
            else:
                # No duplicates - all imported successfully
                return jsonify({
                    'success': True,
                    'has_duplicates': False,
                    'imported_count': imported_count,
                    'message': 'ファイルが正常にアップロードされました'
                })
        else:
            return jsonify({'error': 'ファイルのインポートに失敗しました'}), 500

    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': f'ファイルのアップロードに失敗しました: {str(e)}'}), 500

@app.route('/handle_duplicate_project', methods=['POST'])
def handle_duplicate_project():
    """Handle actions for duplicate projects: skip, update, or import_new."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        action = data.get('action')  # 'skip', 'update', or 'import_new'
        project_data = data.get('project_data')
        existing_id = data.get('existing_id')
        differences = data.get('differences', {})

        if not action or not project_data:
            return jsonify({'error': 'Missing action or project_data'}), 400

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        if action == 'skip':
            # Do nothing, just return success
            conn.close()
            return jsonify({'success': True, 'action': 'skipped'})

        elif action == 'update':
            # Update existing project with only changed fields
            if not existing_id:
                return jsonify({'error': 'Missing existing_id for update'}), 400

            # Build UPDATE statement only for fields that have differences
            set_clause_parts = []
            values = []
            
            if differences:
                # Only update fields that are different
                for field in differences.keys():
                    if field in project_data:
                        set_clause_parts.append(f'"{field}" = ?')
                        values.append(project_data[field])
                
                if not set_clause_parts:
                    # No actual changes to make
                    conn.close()
                    return jsonify({'success': True, 'action': 'no_changes', 'project_id': existing_id})
            else:
                # Fallback: update all fields if differences not provided
                for key, value in project_data.items():
                    if key not in ['id']:  # Skip id field
                        set_clause_parts.append(f'"{key}" = ?')
                        values.append(value)
            
            set_clause = ', '.join(set_clause_parts)
            values.append(existing_id)

            sql = f'''
                UPDATE projects
                SET {set_clause}
                WHERE id = ?
            '''
            cursor.execute(sql, values)
            
            # Add or update history entry for Excel update
            if differences:
                changed_fields = list(differences.keys())
                display_fields = changed_fields[:5]
                if len(changed_fields) > 5:
                    action_details = ', '.join(display_fields) + f' 他{len(changed_fields) - 5}項目'
                else:
                    action_details = ', '.join(display_fields)
                
                # Check if history exists and update or insert
                cursor.execute('SELECT id FROM project_history WHERE project_id = ?', (existing_id,))
                existing_history = cursor.fetchone()
                
                if existing_history:
                    cursor.execute('''
                        UPDATE project_history
                        SET action_type = ?, action_details = ?, created_at = datetime('now', 'localtime')
                        WHERE project_id = ?
                    ''', ('excel_updated', action_details, existing_id))
                else:
                    cursor.execute('''
                        INSERT INTO project_history (project_id, action_type, action_details, created_at)
                        VALUES (?, 'excel_updated', ?, datetime('now', 'localtime'))
                    ''', (existing_id, action_details))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'action': 'updated', 'project_id': existing_id})

        elif action == 'import_new':
            # Insert as new project
            cursor.execute('''
                INSERT INTO projects (
                    SE, "SE(sub)", 案件名, PH, "開発工数（h）", "設計工数（h）", 要件引継, 設計開始,
                    設計完了, 設計書送付, 開発開始, 開発完了, SE納品, BSE, 案件番号, "PJNo.",
                    備考, テスト開始日, テスト完了日, FB完了予定日, ページ数, タスク, ステータス,
                    不要, 注文設計, 注文テスト, 注文FB, 注文BrSE, user_edited_status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                project_data.get('SE', ''),
                project_data.get('SE(sub)', ''),
                project_data.get('案件名', ''),
                project_data.get('PH', ''),
                project_data.get('開発工数（h）', None),
                project_data.get('設計工数（h）', None),
                project_data.get('要件引継', ''),
                project_data.get('設計開始', ''),
                project_data.get('設計完了', ''),
                project_data.get('設計書送付', ''),
                project_data.get('開発開始', ''),
                project_data.get('開発完了', ''),
                project_data.get('SE納品', ''),
                project_data.get('BSE', ''),
                project_data.get('案件番号', ''),
                project_data.get('PJNo.', ''),
                project_data.get('備考', ''),
                project_data.get('テスト開始日', ''),
                project_data.get('テスト完了日', ''),
                project_data.get('FB完了予定日', ''),
                project_data.get('ページ数', None),
                project_data.get('タスク', ''),
                project_data.get('ステータス', '要件引継待ち'),
                project_data.get('不要', 0),
                project_data.get('注文設計', 0),
                project_data.get('注文テスト', 0),
                project_data.get('注文FB', 0),
                project_data.get('注文BrSE', 0),
                project_data.get('user_edited_status', 0)
            ))
            new_id = cursor.lastrowid
            
            # Add history entry for new project from duplicate modal
            cursor.execute('''
                INSERT INTO project_history (project_id, action_type, action_details, created_at)
                VALUES (?, 'created', 'Excelから新規作成 (重複対応)', datetime('now', 'localtime'))
            ''', (new_id,))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'action': 'imported_new', 'project_id': new_id})

        else:
            conn.close()
            return jsonify({'error': 'Invalid action'}), 400

    except Exception as e:
        logging.error(f"Error handling duplicate project: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/upload_mail_template', methods=['POST'])
def upload_mail_template():
    """Handle mail template file upload."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    if not file.filename.endswith('.txt'):
        return jsonify({'error': 'テキストファイル（.txt）のみアップロード可能です'}), 400

    try:
        if not os.path.exists(MAIL_DIR):
            os.makedirs(MAIL_DIR)
        if not os.path.exists(OLD_DIR):
            os.makedirs(OLD_DIR)

        file_path = os.path.join(MAIL_DIR, file.filename)
        if os.path.exists(file_path):
            old_file_path = os.path.join(
                OLD_DIR,
                f"{os.path.splitext(file.filename)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(file.filename)[1]}"
            )
            shutil.move(file_path, old_file_path)
            #logging.info(f"Moved existing file to: {old_file_path}")

        file.save(file_path)
        #logging.info(f"Uploaded new mail template: {file_path}")
        return jsonify({'success': True})

    except Exception as e:
        #logging.error(f"Error uploading mail template: {e}")
        return jsonify({'error': f'ファイルのアップロードに失敗しました: {str(e)}'}), 500




@app.route('/calculate_test_dates', methods=['POST'])
def calculate_test_dates():
    try:
        data = request.get_json()
        page_count = data.get('page_count')
        test_start_date = data.get('test_start_date')
        test_completion_date = data.get('test_completion_date')

        logging.debug(f"Received calculate_test_dates request: page_count={page_count}, test_start_date={test_start_date}, test_completion_date={test_completion_date}")

        # Nếu chỉ có test_completion_date, tính fb_completion_date
        if test_completion_date and not (page_count or test_start_date):
            fb_completion_date = calculate_fb_completion_date(test_completion_date)
            if not fb_completion_date:
                logging.error("Failed to calculate FB completion date")
                return jsonify({'error': 'Failed to calculate FB completion date'}), 400
            logging.debug(f"Calculated fb_completion_date: {fb_completion_date}")
            return jsonify({
                'test_completion_date': test_completion_date,
                'fb_completion_date': fb_completion_date
            })

        # Logic hiện tại: tính cả test_completion_date và fb_completion_date
        if not page_count or not test_start_date:
            logging.error("Missing page_count or test_start_date")
            return jsonify({'error': 'Missing page_count or test_start_date'}), 400

        page_count = int(page_count)
        test_start_date = datetime.strptime(test_start_date, '%Y-%m-%d')

        logging.debug(f"Parsed inputs: page_count={page_count}, test_start_date={test_start_date}")

        test_completion_date = calculate_test_completion_date(page_count, test_start_date.strftime('%Y-%m-%d'))
        fb_completion_date = calculate_fb_completion_date(test_completion_date)

        logging.debug(f"Calculated dates: test_completion_date={test_completion_date}, fb_completion_date={fb_completion_date}")

        return jsonify({
            'test_completion_date': test_completion_date,
            'fb_completion_date': fb_completion_date
        })
    except Exception as e:
        logging.error(f"Error calculating test dates: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/sort_projects', methods=['POST'])
def sort_projects():
    """Sort and filter projects based on column, direction, and search criteria."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        column = data.get('column')
        direction = data.get('direction', 'asc').lower()
        show_all = data.get('show_all', False)
        show_unnecessary = data.get('show_unnecessary', False)
        search_project_name = data.get('search_project_name', '').strip()  # Lấy giá trị tìm kiếm 案件名

        valid_columns = DISPLAY_COLUMNS + ['id', '設計実績', 'テスト実績', 'FB実績', 'BrSE実績']
        if column not in valid_columns:
            return jsonify({'error': 'Invalid column'}), 400
        if direction not in ['asc', 'desc']:
            return jsonify({'error': 'Invalid direction'}), 400

        df = read_projects()
        current_date = datetime.now()
        filtered_projects = []

        for _, row in df.iterrows():
            # Áp dụng bộ lọc: include project if (不要 = 0 OR show_unnecessary) AND (SE納品 not past OR show_all)
            if not show_unnecessary and row.get('不要', 0) == 1:
                continue

            # se_delivery_date = parse_date_from_db(row['SE納品'])
            # if not show_all and se_delivery_date and se_delivery_date.date() < current_date.date():
            #     continue

            # Lọc theo 案件名
            if search_project_name and search_project_name.lower() not in str(row['案件名']).lower():
                continue

            project = row.to_dict()
            project['ステータス'] = calculate_status(project, current_date)

            closest_date = None
            min_diff = float('inf')
            for col in DATE_COLUMNS_DISPLAY:
                date_str_db = row[col]
                date_obj = parse_date_from_db(date_str_db)
                project[f'{col}_past'] = False
                if date_obj is not None:
                    diff = (date_obj.date() - current_date.date()).days
                    if diff >= 0 and diff < min_diff:
                        min_diff = diff
                        closest_date = col
                    if date_obj.date() < current_date.date():
                        project[f'{col}_past'] = True
                project[col] = format_date_jp(date_obj)
            project['highlight_column'] = closest_date if closest_date else None
            
            fb_completion_date = parse_date_from_db(row['FB完了予定日'])
            project['fb_late'] = False

            se_delivery_date = parse_date_from_db(row['SE納品'])
            if fb_completion_date and se_delivery_date:
                project['fb_late'] = fb_completion_date.date() > se_delivery_date.date()

            project = convert_nat_to_none(project)
            filtered_projects.append(project)

        def safe_get(project, key):
            value = project.get(key, 0 if key in ['設計実績', 'テスト実績', 'FB実績', 'BrSE実績'] else '')
            if key in DATE_COLUMNS_DB:
                date_obj = parse_date_for_comparison(value)
                return date_obj if date_obj else datetime.max if direction == 'asc' else datetime.min
            if isinstance(value, str) and value == '':
                return '' if direction == 'asc' else '\uffff'
            return value

        filtered_projects.sort(
            key=lambda x: safe_get(x, column),
            reverse=(direction == 'desc')
        )

        return jsonify({'projects': filtered_projects})
    except Exception as e:
        #logging.error(f"Error sorting projects: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_daily_report_data', methods=['GET'])
def get_daily_report_data():
    """Lấy dữ liệu cho modal báo cáo hàng ngày."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    week_start = request.args.get('week_start')
    if not week_start:
        return jsonify({'error': 'Missing week_start parameter'}), 400

    try:
        datetime.strptime(week_start, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Invalid week_start format'}), 400

    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        cursor.execute('SELECT id, 案件名, "PJNo.", PH FROM projects WHERE 不要 = 0')
        projects = [{'id': row[0], '案件名': row[1], 'PJNo.': row[2], 'PH': row[3]} for row in cursor.fetchall()]

        week_dates = get_week_dates(week_start)
        date_range = [d['date'] for d in week_dates]
        cursor.execute('''
            SELECT project_id, date, task_type, hours
            FROM daily_hours
            WHERE date IN ({})
        '''.format(','.join('?' * len(date_range))), date_range)
        hours = [{'project_id': row[0], 'date': row[1], 'task_type': row[2], 'hours': row[3]} for row in cursor.fetchall()]

        conn.close()
        return jsonify({
            'projects': projects,
            'week_dates': week_dates,
            'hours': hours
        })
    except sqlite3.Error as e:
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/save_daily_hours', methods=['POST'])
def save_daily_hours():
    """Lưu số giờ làm việc hàng ngày vào cơ sở dữ liệu."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    hours_data = data.get('hours', [])

    if not hours_data:
        return jsonify({'success': True})

    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        for entry in hours_data:
            cursor.execute('''
                DELETE FROM daily_hours
                WHERE project_id = ? AND date = ? AND task_type = ?
            ''', (entry['project_id'], entry['date'], entry['task_type']))

        for entry in hours_data:
            cursor.execute('''
                INSERT INTO daily_hours (project_id, date, task_type, hours)
                VALUES (?, ?, ?, ?)
            ''', (entry['project_id'], entry['date'], entry['task_type'], entry['hours']))

        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except sqlite3.Error as e:
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/delete_all_data', methods=['POST'])
def delete_all_data():
    data = request.get_json()
    password = data.get('password', '')
    if 'username' not in session:
        return jsonify({'success': False, 'error': '未認証です'})
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        # Lấy danh sách project_id cần xóa
        cursor.execute('SELECT id FROM projects WHERE 不要 = 1 OR 不要 = "1" OR 不要 = "true"')
        project_ids = [row[0] for row in cursor.fetchall()]
        # Xóa dữ liệu liên quan ở các bảng khác
        for pid in project_ids:
            cursor.execute('DELETE FROM daily_hours WHERE project_id = ?', (pid,))
            cursor.execute('DELETE FROM schedule_done_status WHERE project_id = ?', (pid,))
            cursor.execute('DELETE FROM copied_templates WHERE project_id = ?', (pid,))
        # Xóa project
        cursor.execute('DELETE FROM projects WHERE 不要 = 1 OR 不要 = "1" OR 不要 = "true"')
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/logout')
def logout():
    """Handle user logout."""
    session.pop('username', None)
    flash('ログアウトしました', 'success')
    return redirect(url_for('login'))

def get_temp_dir(project_id):
    """Get the temporary directory path for a project."""
    return os.path.join('temp', str(project_id))

def get_replaced_dir(project_id):
    """Get the replaced files directory path for a project."""
    return os.path.join('replaced', str(project_id))

def read_placeholders(file_path='placeholders.txt'):
    """Read placeholder mappings from a text file, supporting context-dependent rules."""
    placeholders = {'simple': {}, 'context': []}
    try:
        encoding = detect_file_encoding(file_path)
        logging.debug(f"Detected encoding for {file_path}: {encoding}")
        with open(file_path, 'r', encoding=encoding) as f:
            for line_number, line in enumerate(f, 1):
                line = line.strip()
                if not line or '=' not in line:
                    logging.debug(f"Skipping empty or invalid line {line_number}: {line}")
                    continue
                try:
                    # Split on the first '=' to handle values containing '='
                    key, value = map(str.strip, line.split('=', 1))
                    if not key or not value:
                        logging.warning(f"Empty key or value at line {line_number}: {line}. Skipping.")
                        continue
                    # Check for context-dependent rule: [context|target]
                    if key.startswith('[') and key.endswith(']') and '|' in key[1:-1]:
                        try:
                            context, target = key[1:-1].split('|', 1)
                            if not context.strip() or not target.strip():
                                logging.warning(f"Empty context or target at line {line_number}: {line}. Skipping.")
                                continue
                            placeholders['context'].append({
                                'context': context.strip(),
                                'target': target.strip(),
                                'replacement': value.strip()
                            })
                        except ValueError:
                            logging.warning(f"Malformed context-dependent rule at line {line_number}: {line}. Skipping.")
                            continue
                    else:
                        # Handle as simple replacement, removing brackets if present
                        clean_key = key[1:-1].strip() if key.startswith('[') and key.endswith(']') else key.strip()
                        placeholders['simple'][clean_key] = value.strip()
                except Exception as e:
                    logging.warning(f"Error parsing line {line_number}: {line}. Skipping. Error: {str(e)}")
                    continue
        if not placeholders['simple'] and not placeholders['context']:
            logging.error(f"No valid rules found in {file_path}")
            return None
        logging.debug(f"Parsed placeholders: {placeholders}")
        return placeholders
    except UnicodeDecodeError as e:
        logging.error(f"Failed to decode {file_path} with {encoding}: {str(e)}")
        return None
    except FileNotFoundError:
        logging.error(f"Placeholder file {file_path} not found")
        return None
    except Exception as e:
        logging.error(f"Error reading placeholder file {file_path}: {str(e)}")
        return None

@app.route('/upload_mail_edit_files', methods=['POST'])
def upload_mail_edit_files():
    """Handle multiple text file uploads for mail editing."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    if 'files' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    files = request.files.getlist('files')
    if not files or all(file.filename == '' for file in files):
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    temp_dir = 'temp'
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    filenames = []
    try:
        for file in files:
            if not file.filename.endswith('.txt'):
                return jsonify({'error': 'テキストファイル（.txt）のみアップロード可能です'}), 400
            file_path = os.path.join(temp_dir, file.filename)
            file.save(file_path)
            filenames.append(file.filename)
        return jsonify({'success': True, 'filenames': filenames})
    except Exception as e:
        return jsonify({'error': f'ファイルのアップロードに失敗しました: {str(e)}'}), 500

def detect_file_encoding(file_path):
    """Detect the encoding of a file, prioritizing UTF-8 or Shift-JIS."""
    try:
        with open(file_path, 'rb') as f:
            raw_data = f.read()
        result = detect(raw_data)
        encoding = result.get('encoding', 'shift-jis')  # Default to Shift-JIS
        confidence = result.get('confidence', 0)
        # Restrict to UTF-8 or Shift-JIS, fallback to Shift-JIS if uncertain
        if encoding not in ['utf-8', 'shift-jis'] or confidence < 0.8:
            return 'shift-jis'
        return encoding
    except Exception as e:
        logging.error(f"Error detecting encoding for {file_path}: {str(e)}")
        return 'shift-jis'

@app.route('/replace_mail_content', methods=['POST'])
def replace_mail_content():
    """Replace strings in a specific uploaded text file (UTF-8 or Shift-JIS) and output in Shift-JIS, falling back to UTF-8 if needed."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    filename = data.get('filename')
    if not filename:
        return jsonify({'error': 'Missing filename'}), 400

    temp_dir = 'temp'
    replaced_dir = 'replaced'

    # Check if temp directory and file exist
    file_path = os.path.join(temp_dir, filename)
    if not os.path.exists(temp_dir) or not os.path.exists(file_path) or not filename.endswith('.txt'):
        return jsonify({'error': f'ファイル {filename} が見つかりません'}), 404

    # Create or clear replaced directory
    if os.path.exists(replaced_dir):
        shutil.rmtree(replaced_dir)
    os.makedirs(replaced_dir)

    # Read placeholders from file
    placeholders = read_placeholders()
    if not placeholders:
        return jsonify({'error': 'プレースホルダーファイルの読み込みに失敗しました。ファイルのエンコーディングを確認してください。'}), 500

    results = []
    failed_files = []
    try:
        # Detect file encoding
        encoding = detect_file_encoding(file_path)
        logging.debug(f"Detected encoding for {filename}: {encoding}")
        with open(file_path, 'r', encoding=encoding) as f:
            content = f.read()

        # Apply context-dependent replacements
        for rule in placeholders['context']:
            try:
                if not isinstance(rule, dict) or 'context' not in rule or 'target' not in rule or 'replacement' not in rule:
                    logging.error(f"Invalid context-dependent rule: {rule}. Skipping.")
                    failed_files.append(f"{filename}: 無効なコンテキスト依存ルールが見つかりました。placeholders.txtを確認してください。")
                    continue
                context = re.escape(rule['context'])
                target = re.escape(rule['target'])
                pattern = f'{context}\n{target}'
                replacement = f"{rule['context']}\n{rule['replacement']}"
                content = re.sub(pattern, replacement, content)
            except Exception as e:
                logging.error(f"Error applying context-dependent rule {rule}: {str(e)}")
                failed_files.append(f"{filename}: コンテキスト依存ルールの適用中にエラーが発生しました: {str(e)}")
                continue

        # Apply simple replacements
        for key, value in placeholders['simple'].items():
            content = content.replace(key, value)

        replaced_file_path = os.path.join(replaced_dir, filename)
        # Try writing in Shift-JIS first
        try:
            with open(replaced_file_path, 'w', encoding='shift-jis') as f:
                f.write(content)
            output_encoding = 'shift-jis'
        except UnicodeEncodeError as e:
            # Fallback to UTF-8 if Shift-JIS fails
            logging.warning(f"Failed to encode {filename} in Shift-JIS: {str(e)}. Falling back to UTF-8.")
            with open(replaced_file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            output_encoding = 'utf-8'
            failed_files.append(
                f"{filename}: Shift-JISでサポートされていない文字（例: ①）が含まれています。UTF-8で出力されました。"
            )

        # Convert content to UTF-8 for JSON response (for UI display)
        results.append(f"ファイル: {filename} (出力エンコーディング: {output_encoding})\n{content.encode('utf-8', errors='replace').decode('utf-8', errors='replace')}")

        if failed_files:
            error_message = "ファイルで問題が発生しました:\n" + "\n".join(failed_files)
            return jsonify({'success': True, 'results': results, 'warnings': error_message})

        return jsonify({'success': True, 'results': results})
    except UnicodeDecodeError as e:
        logging.error(f"Failed to decode {filename} with {encoding}: {str(e)}")
        return jsonify({'error': f"{filename}: エンコーディングが無効です（UTF-8またはShift-JISを期待）"}), 500
    except Exception as e:
        logging.error(f"Error processing {filename}: {str(e)}")
        return jsonify({'error': f"{filename}: 処理中にエラーが発生しました: {str(e)}"}), 500

@app.route('/download_replaced_files', methods=['GET'])
def download_replaced_files():
    """Download replaced files (Shift-JIS) as a ZIP archive."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    replaced_dir = 'replaced'
    if not os.path.exists(replaced_dir):
        return jsonify({'error': 'リプレースされたファイルが見つかりません'}), 404

    try:
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename in os.listdir(replaced_dir):
                if filename.endswith('.txt'):
                    file_path = os.path.join(replaced_dir, filename)
                    zf.write(file_path, filename)
        memory_file.seek(0)

        # Clean up temp and replaced directories
        temp_dir = 'temp'
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        if os.path.exists(replaced_dir):
            shutil.rmtree(replaced_dir)

        return app.response_class(
            memory_file,
            mimetype='application/zip',
            headers={'Content-Disposition': 'attachment; filename=replaced_mail_templates.zip'}
        )
    except Exception as e:
        return jsonify({'error': f'ダウンロードに失敗しました: {str(e)}'}), 500

@app.route('/get_mail_template_list', methods=['GET'])
def get_mail_template_list():
    """Lấy danh sách các file txt trong thư mục mail."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        templates = get_mail_templates()
        return jsonify({'templates': [full_name for full_name, _ in templates]})
    except Exception as e:
        logging.error(f"Error fetching mail template list: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/save_replaced_file', methods=['POST'])
def save_replaced_file():
    """Lưu nội dung đã replace vào thư mục mail với tên file được chỉ định."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    filename = data.get('filename')
    content = data.get('content')
    overwrite = data.get('overwrite', False)

    if not filename or not content:
        return jsonify({'error': 'Missing filename or content'}), 400

    if not filename.endswith('.txt'):
        filename += '.txt'

    file_path = os.path.join(MAIL_DIR, filename)
    try:
        if os.path.exists(file_path) and not overwrite:
            return jsonify({'error': f'File {filename} already exists. Use overwrite option.'}), 400

        if not os.path.exists(MAIL_DIR):
            os.makedirs(MAIL_DIR)

        # Viết file với encoding Shift-JIS, fallback sang UTF-8 nếu cần
        try:
            with open(file_path, 'w', encoding='shift-jis') as f:
                f.write(content)
            encoding = 'shift-jis'
        except UnicodeEncodeError:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            encoding = 'utf-8'

        return jsonify({'success': True, 'encoding': encoding})
    except Exception as e:
        logging.error(f"Error saving replaced file {filename}: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/clear_temp_files', methods=['POST'])
def clear_temp_files():
    """Xóa tất cả file trong thư mục temp và replaced."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    temp_dir = 'temp'
    replaced_dir = 'replaced'
    try:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        if os.path.exists(replaced_dir):
            shutil.rmtree(replaced_dir)
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error clearing temp files: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/get_template_content/<filename>', methods=['GET'])
def get_template_content(filename):
    """Get content of a template file for editing."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    # Security check for filename
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        return jsonify({'error': 'Invalid filename'}), 400
    
    if not filename.endswith('.txt'):
        return jsonify({'error': 'Invalid file type'}), 400
    
    file_path = os.path.join(MAIL_DIR, filename)
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        encoding = detect_file_encoding(file_path)
        with open(file_path, 'r', encoding=encoding) as f:
            content = f.read()
        return jsonify({'success': True, 'content': content, 'encoding': encoding})
    except Exception as e:
        logging.error(f"Error reading template file {filename}: {str(e)}")
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 500


@app.route('/save_template_content', methods=['POST'])
def save_template_content():
    """Save edited template content."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    filename = data.get('filename')
    content = data.get('content')
    
    if not filename or content is None:
        return jsonify({'error': 'Missing filename or content'}), 400
    
    # Security check for filename
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        return jsonify({'error': 'Invalid filename'}), 400
    
    if not filename.endswith('.txt'):
        return jsonify({'error': 'Invalid file type'}), 400
    
    file_path = os.path.join(MAIL_DIR, filename)
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        # Try to save with Shift-JIS encoding, fallback to UTF-8
        try:
            with open(file_path, 'w', encoding='shift-jis') as f:
                f.write(content)
            encoding = 'shift-jis'
        except UnicodeEncodeError:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            encoding = 'utf-8'
        
        return jsonify({'success': True, 'encoding': encoding})
    except Exception as e:
        logging.error(f"Error saving template file {filename}: {str(e)}")
        return jsonify({'error': f'Failed to save file: {str(e)}'}), 500


@app.route('/delete_mail_template', methods=['POST'])
def delete_mail_template():
    """Delete a mail template file from the mail directory."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    filename = data.get('filename')

    print(f"file name: {filename}")

    if not filename:
        logging.error("Missing filename in delete_mail_template request")
        return jsonify({'error': 'Missing filename'}), 400

    if not filename.endswith('.txt'):
        logging.error(f"Invalid file extension for {filename}: Only .txt files are allowed")
        return jsonify({'error': 'Only .txt files are allowed'}), 400

    # Prevent path traversal attacks
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        logging.error(f"Invalid filename detected: {filename}")
        return jsonify({'error': 'Invalid filename'}), 400

    file_path = os.path.join(MAIL_DIR, filename)
    print(f"file path: {file_path}")

    if not os.path.exists(file_path):
        print("go 4")
        logging.error(f"File not found: {file_path}")
        return jsonify({'error': 'File not found'}), 404

    try:
        print("go 1")
        os.remove(file_path)
        print("go 2")
        logging.info(f"Successfully deleted mail template: {file_path}")
        return jsonify({'success': True})
    except Exception as e:
        print("go 3")
        logging.error(f"Error deleting mail template {filename}: {str(e)}")
        return jsonify({'error': f'Failed to delete file: {str(e)}'}), 500

@app.route('/rename_mail_template', methods=['POST'])
def rename_mail_template():
    """Rename a mail template file in the mail directory."""
    if 'username' not in session:
        logging.error("Unauthorized access to /rename_mail_template")
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    old_filename = data.get('old_filename')
    new_filename = data.get('new_filename')

    logging.info(f"Received rename request: old_filename={old_filename}, new_filename={new_filename}")

    if not old_filename or not new_filename:
        logging.error("Missing old_filename or new_filename")
        return jsonify({'error': 'Old and new filenames are required'}), 400

    old_filepath = os.path.join(MAIL_DIR, old_filename)
    new_filepath = os.path.join(MAIL_DIR, new_filename)

    logging.info(f"Attempting to rename {old_filepath} to {new_filepath}")

    try:
        if not os.path.exists(MAIL_DIR):
            logging.error(f"Mail directory does not exist: {MAIL_DIR}")
            return jsonify({'error': f'Mail directory does not exist'}), 500

        if not os.path.exists(old_filepath):
            logging.error(f"Source file does not exist: {old_filepath}")
            return jsonify({'error': f'File {old_filename} does not exist'}), 404

        if os.path.exists(new_filepath):
            logging.error(f"Destination file already exists: {new_filepath}")
            return jsonify({'error': f'File {new_filename} already exists'}), 400

        os.rename(old_filepath, new_filepath)
        logging.info(f"Successfully renamed {old_filename} to {new_filename}")
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error renaming mail template: {str(e)}")
        return jsonify({'error': str(e)}), 500

def validate_email(email):
    """Validate email format."""
    if not email:
        return True  # Allow empty emails
    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    return bool(re.match(email_regex, email))

@app.route('/get_email_data', methods=['GET'])
def get_email_data():
    """Fetch SE and SE(sub) email addresses and manager info."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        # Fetch unique SE and SE(sub) names from projects table
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT SE FROM projects WHERE SE IS NOT NULL AND SE != ""')
        se_names = [row[0] for row in cursor.fetchall()]
        cursor.execute('SELECT DISTINCT "SE(sub)" FROM projects WHERE "SE(sub)" IS NOT NULL AND "SE(sub)" != ""')
        se_sub_names = [row[0] for row in cursor.fetchall()]
        conn.close()

        # Kết hợp SE và SE(sub)
        all_se_names = list(set(se_names + se_sub_names))

        # Read SE_email.csv
        se_email_file = os.path.join(MAIL_DIR, 'SE_email.csv')
        se_emails = []
        if os.path.exists(se_email_file):
            df = pd.read_csv(se_email_file, encoding='utf-8')
            # Replace NaN with empty string and ensure all values are strings
            df = df.fillna('')
            df['SE'] = df['SE'].astype(str)
            df['email'] = df['email'].astype(str)
            # Replace 'nan' string with empty string
            df['email'] = df['email'].replace('nan', '')
            
            existing_se = set(df['SE'].tolist())
            for se in all_se_names:
                if se in existing_se:
                    email = df[df['SE'] == se]['email'].iloc[0]
                    # Ensure email is string and not NaN
                    email = str(email) if pd.notna(email) else ''
                    if email == 'nan':
                        email = ''
                else:
                    email = ''
                se_emails.append({'SE': str(se), 'email': email})
        else:
            se_emails = [{'SE': str(se), 'email': ''} for se in all_se_names]

        # Read kanrisha.csv
        kanrisha_file = os.path.join(MAIL_DIR, 'kanrisha.csv')
        manager = None
        if os.path.exists(kanrisha_file):
            df = pd.read_csv(kanrisha_file, encoding='utf-8')
            if not df.empty:
                # Handle NaN values in manager data
                name = df['name'].iloc[0] if pd.notna(df['name'].iloc[0]) else ''
                email = df['email'].iloc[0] if pd.notna(df['email'].iloc[0]) else ''
                # Convert to string and handle 'nan' string
                name = str(name) if name != 'nan' else ''
                email = str(email) if email != 'nan' else ''
                manager = {'name': name, 'email': email}

        return jsonify({'se_emails': se_emails, 'manager': manager})
    except Exception as e:
        logging.error(f"Error fetching email data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/save_email_data', methods=['POST'])
def save_email_data():
    """Save SE email addresses and manager info to CSV files."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        se_emails = data.get('se_emails', [])
        manager = data.get('manager', {})

        # Validate SE emails
        for se in se_emails:
            if not validate_email(se.get('email')):
                return jsonify({'error': f"無効なメールアドレスです: {se.get('SE')} のメールアドレスを確認してください"}), 400

        # Validate manager email
        if manager.get('email') and not validate_email(manager.get('email')):
            return jsonify({'error': '無効な管理者メールアドレスです。メールアドレスを確認してください'}), 400

        # Ensure MAIL_DIR exists
        if not os.path.exists(MAIL_DIR):
            os.makedirs(MAIL_DIR)

        # Save SE_email.csv
        se_email_file = os.path.join(MAIL_DIR, 'SE_email.csv')
        df_se = pd.DataFrame(se_emails)
        df_se.to_csv(se_email_file, index=False, encoding='utf-8')

        # Save kanrisha.csv
        kanrisha_file = os.path.join(MAIL_DIR, 'kanrisha.csv')
        df_manager = pd.DataFrame([manager] if manager else [], columns=['name', 'email'])
        df_manager.to_csv(kanrisha_file, index=False, encoding='utf-8')

        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error saving email data: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Hàm hỗ trợ
def detect_file_encoding(file_path):
    """Detect file encoding using chardet."""
    with open(file_path, 'rb') as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        return result['encoding'] or 'utf-8'

def parse_date_from_db(date_str):
    """Parse date string from database."""
    if not date_str or date_str == '':
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None

def format_date_jp(date_obj):
    """Format date to Japanese format (YYYY/MM/DD(曜日))."""
    if date_obj is None:
        return ''
    weekdays = ['月', '火', '水', '木', '金', '土', '日']
    weekday = weekdays[date_obj.weekday()]
    return date_obj.strftime('%Y/%m/%d') + f'({weekday})'

def convert_nat_to_none(project_dict):
    """Convert NaT or None-like values to None."""
    for key, value in project_dict.items():
        if value is None or value != value:  # NaN/NaT check
            project_dict[key] = ''
    return project_dict

def read_se_emails():
    """Read SE and SE(sub) emails from mail/SE_email.csv."""
    se_emails = {}
    file_path = os.path.join(MAIL_DIR, 'SE_email.csv')
    if not os.path.exists(file_path):
        return se_emails

    encoding = detect_file_encoding(file_path)
    try:
        with open(file_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return se_emails
            for row in reader:
                if row['email']:
                    se_emails[row['SE']] = row['email']
                else:
                    se_emails[row['SE']] = row['SE']
        return se_emails
    except Exception as e:
        #logging.error(f"Failed to read SE_email.csv: {e}")
        return se_emails

def read_manager_email():
    """Read manager email from mail/kanrisha.csv."""
    file_path = os.path.join(MAIL_DIR, 'kanrisha.csv')
    if not os.path.exists(file_path):
        return ''

    encoding = detect_file_encoding(file_path)
    try:
        with open(file_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:  # Kiểm tra nếu file rỗng
                return ''
            for row in reader:
                return row['email'] if row['email'] else row['name']  # Trả về email đầu tiên, nếu có
        return ''
    except Exception as e:
        #logging.error(f"Failed to read kanrisha.csv: {e}")
        return ''

@app.route('/get_copied_templates/<int:project_id>', methods=['GET'])
def get_copied_templates(project_id):
    """Get list of copied templates for a project."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT filename FROM copied_templates
            WHERE project_id = ?
        ''', (project_id,))
        templates = [row[0] for row in cursor.fetchall()]
        conn.close()
        #logging.debug(f"Copied templates for project_id={project_id}: {templates}")
        return jsonify({'templates': templates})
    except sqlite3.Error as e:
        #logging.error(f"Database error while fetching copied templates: {e}")
        return jsonify({'error': 'Database error'}), 500

@app.route('/get_mail_content/<int:project_id>/<filename>', methods=['GET'])
def get_mail_content(project_id, filename):
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        #logging.error(f"Invalid filename detected: {filename}")
        return jsonify({'error': 'Invalid filename'}), 400
    file_path = os.path.join(MAIL_DIR, filename)
    if not os.path.exists(file_path) or not filename.endswith('.txt'):
        #logging.error(f"File not found or invalid: {file_path}")
        return jsonify({'error': 'File not found or not a .txt file'}), 404

    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        columns = [description[0] for description in cursor.description]
        conn.close()
    except sqlite3.Error as e:
        #logging.error(f"Database error: {e}")
        return jsonify({'error': 'Database error'}), 500

    if not project:
        #logging.error(f"Project not found for ID: {project_id}")
        return jsonify({'error': 'プロジェクトが見つかりません'}), 404

    project_dict = {col: project[i] for i, col in enumerate(columns)}
    project_dict = convert_nat_to_none(project_dict)

    encoding = detect_file_encoding(file_path)
    try:
        with open(file_path, 'r', encoding=encoding) as f:
            content = f.read()
    except Exception as e:
        #logging.error(f"Failed to read file {file_path}: {e}")
        return jsonify({'error': f'ファイルの読み込みに失敗しました: {str(e)}'}), 500

    # Đọc email của SE và SE(sub)
    se_emails = read_se_emails()
    manager_email = read_manager_email()
    se_name = project_dict.get('SE', '')
    se_sub_name = project_dict.get('SE(sub)', '')

    # Chuẩn bị placeholders
    pjno_value = project_dict.get('PJNo.', '')
    if isinstance(pjno_value, (float, int)):
        pjno_value = str(int(pjno_value))
    else:
        pjno_value = str(pjno_value)

    ph = project_dict.get('PH', '')
    if ph != '':
        ph = ' PH' + ph

    placeholders = {
        '{anken_name}': project_dict.get('案件名', ''),
        '{se_name}': se_name,
        '{pj}': pjno_value,
        '{開発工数}': project_dict.get('開発工数（h）', ''),
        '{PH}': ph
    }
    

    # Thêm {se_mail} với định dạng 'email SE, email SE(sub)'
    se_mail_value = ''
    if se_emails and se_name in se_emails and se_emails[se_name]:
        se_mail_value += se_emails[se_name]
    if se_emails and se_sub_name in se_emails and se_emails[se_sub_name]:
        if se_mail_value:
            se_mail_value += ', '
        se_mail_value += se_emails[se_sub_name]
    placeholders['{se_mail}'] = se_mail_value

    # Thêm {mail}: luôn thay bằng '設計チーム <pjpromotion@j-ems.jp>', thêm manager_email nếu có
    mail_value = '設計チーム <pjpromotion@j-ems.jp>'
    if manager_email:
        mail_value += f', {manager_email}'
    placeholders['{mail}'] = mail_value

    for date_col in DATE_COLUMNS_DB:
        date_str = project_dict.get(date_col, '')
        date_obj = parse_date_from_db(date_str)
        placeholders[f'{{{date_col}}}'] = format_date_jp(date_obj)

    if not ph.strip():
    # Xóa các trường hợp như "({PH})", "( {PH} )", "（{PH}）", "（ {PH} ）"
        import re
        content = re.sub(r'[\(（]\s*\{PH\}\s*[\)）]', '', content)
        
    for key, value in placeholders.items():
        content = content.replace(key, str(value))
    return jsonify({'content': content})

@app.route('/save_copied_template', methods=['POST'])
def save_copied_template():
    """Save copied mail template to database."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json()
    project_id = data.get('project_id')
    filename = data.get('filename')
    if not project_id or not filename:
        #logging.error(f"Missing project_id or filename: project_id={project_id}, filename={filename}")
        return jsonify({'error': 'Missing project_id or filename'}), 400
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO copied_templates (project_id, filename, copied_at)
            VALUES (?, ?, ?)
        ''', (project_id, filename, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        
        # Add or update history entry for mail template usage
        # Remove .txt extension for display
        template_name = filename.replace('.txt', '')
        
        # Check if history exists and update or insert
        cursor.execute('SELECT id FROM project_history WHERE project_id = ?', (project_id,))
        existing = cursor.fetchone()
        
        if existing:
            cursor.execute('''
                UPDATE project_history
                SET action_type = ?, action_details = ?, created_at = datetime('now', 'localtime')
                WHERE project_id = ?
            ''', ('mail_sent', template_name, project_id))
        else:
            cursor.execute('''
                INSERT INTO project_history (project_id, action_type, action_details, created_at)
                VALUES (?, 'mail_sent', ?, datetime('now', 'localtime'))
            ''', (project_id, template_name))
        
        conn.commit()
        conn.close()
        #logging.debug(f"Saved copied template: project_id={project_id}, filename={filename}")
        return jsonify({'success': True})
    except sqlite3.Error as e:
        #logging.error(f"Database error while saving copied template: {e}")
        return jsonify({'error': 'Database error'}), 500

@app.route('/remove_copied_template', methods=['POST'])
def remove_copied_template():
    """Remove a copied mail template from the database."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json()
    project_id = data.get('project_id')
    filename = data.get('filename')
    if not project_id or not filename:
        #logging.error(f"Missing project_id or filename: project_id={project_id}, filename={filename}")
        return jsonify({'error': 'Missing project_id or filename'}), 400
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            DELETE FROM copied_templates
            WHERE project_id = ? AND filename = ?
        ''', (project_id, filename))
        if cursor.rowcount == 0:
            conn.close()
            #logging.debug(f"No copied template found: project_id={project_id}, filename={filename}")
            return jsonify({'error': 'Template not found'}), 404
        conn.commit()
        conn.close()
        #logging.debug(f"Removed copied template: project_id={project_id}, filename={filename}")
        return jsonify({'success': True})
    except sqlite3.Error as e:
        conn.close()
        #logging.error(f"Database error while removing copied template: {e}")
        return jsonify({'error': 'Database error'}), 500


@app.route('/get_schedule_data', methods=['GET'])
def get_schedule_data():
    try:
        week_start_str = request.args.get('week_start')
        logging.info(f"Received week_start: {week_start_str}")
        if not week_start_str:
            logging.error("Week start date is missing")
            return jsonify({'error': 'Week start date is required'}), 400

        try:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d')
        except ValueError as ve:
            logging.error(f"Invalid week_start format: {week_start_str}")
            return jsonify({'error': 'Invalid week start date format'}), 400

        week_end = week_start + timedelta(days=6)

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            'SELECT id, 案件名, 要件引継, 設計開始, 設計完了, 設計書送付, 開発開始, 開発完了, テスト開始日, テスト完了日, FB完了予定日, SE納品, SE, PH, "開発工数（h）" FROM projects WHERE 不要 = 0')
        projects = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]

        cursor.execute('SELECT project_id, date_column, done FROM schedule_done_status')
        done_statuses = {f"{row[0]}_{row[1]}": bool(row[2]) for row in cursor.fetchall()}

        conn.close()

        date_columns = ['要件引継', '設計開始', '設計完了', '設計書送付', '開発開始', '開発完了', 'テスト開始日',
                        'テスト完了日', 'FB完了予定日', 'SE納品']
        japanese_days = ['月', '火', '水', '木', '金', '土', '日']
        filtered_projects = []

        for project in projects:
            for date_col in date_columns:
                date_str = project.get(date_col)
                if date_str and isinstance(date_str, str):
                    try:
                        date_obj = parse(date_str)
                        if week_start.date() <= date_obj.date() <= week_end.date():
                            done_key = f"{project['id']}_{date_col}"
                            date_value = date_obj.strftime('%d/%m/%Y')
                            day_name = japanese_days[date_obj.weekday()]
                            filtered_projects.append({
                                'id': project['id'],
                                '案件名': project['案件名'],
                                'date_column': date_col,
                                'date_value': date_value,
                                'day_name': day_name,
                                'schedule_done': done_statuses.get(done_key, False),
                                'SE': project['SE'] or '',
                                'PH': project['PH'] or '',
                                '開発工数（h）': project['開発工数（h）'] if project['開発工数（h）'] is not None else ''
                            })
                    except ValueError:
                        logging.warning(f"Invalid date format for {date_col} in project {project['id']}: {date_str}")
                        continue

        return jsonify({'projects': filtered_projects})
    except Exception as e:
        logging.error(f"Error fetching schedule data: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/save_schedule_done', methods=['POST'])
def save_schedule_done():
    try:
        data = request.get_json()
        projects = data.get('projects', [])

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()

        for project in projects:
            project_id = project.get('id')
            date_column = project.get('date_column')
            done = 1 if project.get('schedule_done') else 0

            # Kiểm tra xem bản ghi đã tồn tại chưa
            cursor.execute('SELECT id FROM schedule_done_status WHERE project_id = ? AND date_column = ?',
                           (project_id, date_column))
            existing = cursor.fetchone()

            if existing:
                # Cập nhật bản ghi
                cursor.execute('UPDATE schedule_done_status SET done = ? WHERE project_id = ? AND date_column = ?',
                               (done, project_id, date_column))
            else:
                # Thêm bản ghi mới
                cursor.execute('INSERT INTO schedule_done_status (project_id, date_column, done) VALUES (?, ?, ?)',
                               (project_id, date_column, done))

        conn.commit()
        conn.close()

        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error saving schedule done status: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/editor_list')
def editor_list():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('editor_list.html')

@app.route('/editor')
def editor():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    document_id = request.args.get('id')
    document = None
    
    if document_id:
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, title, content, created_at, updated_at 
                FROM editor_document 
                WHERE id = ?
            ''', (document_id,))
            row = cursor.fetchone()
            conn.close()
            
            if row:
                document = {
                    'id': row[0],
                    'title': row[1],
                    'content': row[2],
                    'created_at': row[3],
                    'updated_at': row[4]
                }
        except Exception as e:
            logging.error(f"Error loading document: {str(e)}")
    
    return render_template('editor.html', document=document)

@app.route('/api/editor_documents', methods=['GET'])
def get_editor_documents():
    """Get all editor documents."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, title, content, created_at, updated_at 
            FROM editor_document 
            ORDER BY updated_at DESC
        ''')
        documents = []
        for row in cursor.fetchall():
            documents.append({
                'id': row[0],
                'title': row[1],
                'content': row[2],
                'created_at': row[3],
                'updated_at': row[4]
            })
        conn.close()
        return jsonify({'documents': documents})
    except Exception as e:
        logging.error(f"Error fetching editor documents: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/editor_documents', methods=['POST'])
def create_editor_document():
    """Create a new editor document."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
            INSERT INTO editor_document (title, content, created_at, updated_at)
            VALUES (?, ?, ?, ?)
        ''', (title, content, current_time, current_time))
        document_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': document_id})
    except Exception as e:
        logging.error(f"Error creating editor document: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/editor_documents/<int:document_id>', methods=['GET'])
def get_editor_document(document_id):
    """Get a specific editor document."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, title, content, created_at, updated_at 
            FROM editor_document 
            WHERE id = ?
        ''', (document_id,))
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return jsonify({'error': 'Document not found'}), 404
        
        document = {
            'id': row[0],
            'title': row[1],
            'content': row[2],
            'created_at': row[3],
            'updated_at': row[4]
        }
        return jsonify({'document': document})
    except Exception as e:
        logging.error(f"Error fetching editor document: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/editor_documents/<int:document_id>', methods=['PUT'])
def update_editor_document(document_id):
    """Update an existing editor document."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
            UPDATE editor_document 
            SET title = ?, content = ?, updated_at = ?
            WHERE id = ?
        ''', (title, content, current_time, document_id))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Document not found'}), 404
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error updating editor document: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/editor_documents/<int:document_id>', methods=['DELETE'])
def delete_editor_document(document_id):
    """Delete an editor document."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM editor_document WHERE id = ?', (document_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Document not found'}), 404
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error deleting editor document: {str(e)}")
        return jsonify({'error': str(e)}), 500

def compress_file(file_path, original_filename):
    """Compress file based on its type."""
    file_extension = os.path.splitext(original_filename)[1].lower()
    
    # Skip compression for image files
    if file_extension in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
        logging.info(f"Skipping compression for image file: {original_filename}")
        return False
    
    # Skip compression for already compressed files
    if file_extension in ['.zip', '.rar', '.7z', '.gz', '.tar', '.bz2']:
        logging.info(f"Skipping compression for already compressed file: {original_filename}")
        return False
    
    # For other files, use ZIP compression
    try:
        # Get original file size
        original_size = os.path.getsize(file_path)
        logging.info(f"Original file size for {original_filename}: {original_size} bytes")
        
        # Create compressed version
        compressed_path = file_path + '_temp.zip'
        with zipfile.ZipFile(compressed_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            zf.write(file_path, original_filename)
        
        # Check if compression actually reduced file size significantly (at least 10%)
        compressed_size = os.path.getsize(compressed_path)
        logging.info(f"Compressed file size for {original_filename}: {compressed_size} bytes")
        
        if compressed_size < original_size * 0.9:  # Only replace if at least 10% reduction
            # Replace original with compressed
            os.replace(compressed_path, file_path)
            logging.info(f"Successfully compressed {original_filename}: {original_size} -> {compressed_size} bytes ({((original_size - compressed_size) / original_size * 100):.1f}% reduction)")
            return True
        else:
            # Remove compressed file and keep original
            os.remove(compressed_path)
            logging.info(f"Compression not beneficial for {original_filename} ({((original_size - compressed_size) / original_size * 100):.1f}% reduction), keeping original")
            return False
            
    except Exception as e:
        logging.error(f"Error compressing file {original_filename}: {str(e)}")
        # Clean up any temporary file
        compressed_path = file_path + '_temp.zip'
        if os.path.exists(compressed_path):
            try:
                os.remove(compressed_path)
            except:
                pass
        return False
    
@app.route('/memo_list')
def memo_list():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('memo_list.html')

@app.route('/memo')
def memo():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    memo_id = request.args.get('id')
    memo_data = None
    memo_files = []
    
    if memo_id:
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, title, content, created_at, updated_at 
                FROM memo 
                WHERE id = ?
            ''', (memo_id,))
            row = cursor.fetchone()
            
            if row:
                memo_data = {
                    'id': row[0],
                    'title': row[1],
                    'content': row[2],
                    'created_at': row[3],
                    'updated_at': row[4]
                }
                
                # Get memo files
                cursor.execute('''
                    SELECT id, filename, original_filename, file_type, file_size, uploaded_at
                    FROM memo_files 
                    WHERE memo_id = ?
                    ORDER BY uploaded_at DESC
                ''', (memo_id,))
                memo_files = [
                    {
                        'id': file_row[0],
                        'filename': file_row[1],
                        'original_filename': file_row[2],
                        'file_type': file_row[3],
                        'file_size': file_row[4],
                        'uploaded_at': file_row[5]
                    }
                    for file_row in cursor.fetchall()
                ]
            
            conn.close()
        except Exception as e:
            logging.error(f"Error loading memo: {str(e)}")
    
    return render_template('memo.html', memo=memo_data, memo_files=memo_files)

@app.route('/api/memos', methods=['GET'])
def get_memos():
    """Get all memos."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT m.id, m.title, m.content, m.created_at, m.updated_at,
                   COUNT(mf.id) as file_count
            FROM memo m
            LEFT JOIN memo_files mf ON m.id = mf.memo_id
            GROUP BY m.id
            ORDER BY m.updated_at DESC
        ''')
        memos = []
        for row in cursor.fetchall():
            memos.append({
                'id': row[0],
                'title': row[1],
                'content': row[2],
                'created_at': row[3],
                'updated_at': row[4],
                'file_count': row[5]
            })
        conn.close()
        return jsonify({'memos': memos})
    except Exception as e:
        logging.error(f"Error fetching memos: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/memos', methods=['POST'])
def create_memo():
    """Create a new memo."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
            INSERT INTO memo (title, content, created_at, updated_at)
            VALUES (?, ?, ?, ?)
        ''', (title, content, current_time, current_time))
        memo_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': memo_id})
    except Exception as e:
        logging.error(f"Error creating memo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/memos/<int:memo_id>', methods=['GET'])
def get_memo(memo_id):
    """Get a specific memo."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, title, content, created_at, updated_at 
            FROM memo 
            WHERE id = ?
        ''', (memo_id,))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return jsonify({'error': 'Memo not found'}), 404
        
        memo = {
            'id': row[0],
            'title': row[1],
            'content': row[2],
            'created_at': row[3],
            'updated_at': row[4]
        }
        
        # Get memo files
        cursor.execute('''
            SELECT id, filename, original_filename, file_type, file_size, uploaded_at
            FROM memo_files 
            WHERE memo_id = ?
            ORDER BY uploaded_at DESC
        ''', (memo_id,))
        files = [
            {
                'id': file_row[0],
                'filename': file_row[1],
                'original_filename': file_row[2],
                'file_type': file_row[3],
                'file_size': file_row[4],
                'uploaded_at': file_row[5]
            }
            for file_row in cursor.fetchall()
        ]
        
        conn.close()
        return jsonify({'memo': memo, 'files': files})
    except Exception as e:
        logging.error(f"Error fetching memo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/memos/<int:memo_id>', methods=['PUT'])
def update_memo(memo_id):
    """Update an existing memo."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('''
            UPDATE memo 
            SET title = ?, content = ?, updated_at = ?
            WHERE id = ?
        ''', (title, content, current_time, memo_id))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Memo not found'}), 404
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error updating memo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/memos/<int:memo_id>', methods=['DELETE'])
def delete_memo(memo_id):
    """Delete a memo and its associated files."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Get files associated with this memo
        cursor.execute('SELECT filename FROM memo_files WHERE memo_id = ?', (memo_id,))
        files = cursor.fetchall()
        
        # Delete files from filesystem
        for file_row in files:
            file_path = os.path.join(FILEUPLOAD_DIR, file_row[0])
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    logging.error(f"Error deleting file {file_path}: {str(e)}")
        
        # Delete from database
        cursor.execute('DELETE FROM memo_files WHERE memo_id = ?', (memo_id,))
        cursor.execute('DELETE FROM memo WHERE id = ?', (memo_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Memo not found'}), 404
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error deleting memo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/memo_files/upload', methods=['POST'])
def upload_memo_file():
    """Upload files for a memo."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    memo_id = request.form.get('memo_id')
    if not memo_id:
        return jsonify({'error': 'Memo ID is required'}), 400
    
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    if not files or all(file.filename == '' for file in files):
        return jsonify({'error': 'No files selected'}), 400
    
    # Create upload directory if it doesn't exist
    if not os.path.exists(FILEUPLOAD_DIR):
        os.makedirs(FILEUPLOAD_DIR)
    
    uploaded_files = []
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        for file in files:
            if file.filename == '':
                continue
            
            # Generate unique filename
            original_filename = file.filename
            file_extension = os.path.splitext(original_filename)[1].lower()
            unique_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{memo_id}_{len(uploaded_files)}{file_extension}"
            
            # Determine file type
            if file_extension in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
                file_type = 'image'
            else:
                file_type = 'file'
            
            # Save file
            file_path = os.path.join(FILEUPLOAD_DIR, unique_filename)
            file.save(file_path)
            logging.info(f"Saved file: {file_path}")
            
            # Get original file size
            original_size = os.path.getsize(file_path)
            
            # Compress file (only non-image files)
            logging.info(f"Attempting to compress file: {original_filename}, type: {file_type}")
            was_compressed = compress_file(file_path, original_filename)
            logging.info(f"Compression result for {original_filename}: {was_compressed}")
            
            # Update filename if compressed
            stored_filename = unique_filename
            if was_compressed:
                # File was compressed, so it's now a ZIP file
                base_name = os.path.splitext(unique_filename)[0]
                stored_filename = f"{base_name}.zip"
                # Rename the file in filesystem to reflect the ZIP extension
                new_file_path = os.path.join(FILEUPLOAD_DIR, stored_filename)
                os.rename(file_path, new_file_path)
                file_path = new_file_path
                logging.info(f"Renamed compressed file to: {stored_filename}")
            
            # Get final file size
            final_size = os.path.getsize(file_path)
            
            # Calculate compression ratio
            compression_ratio = None
            if was_compressed and original_size > 0:
                compression_ratio = f"{((original_size - final_size) / original_size * 100):.1f}%"
            
            # Save to database with correct filename
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('''
                INSERT INTO memo_files (memo_id, filename, original_filename, file_type, file_size, uploaded_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (memo_id, stored_filename, original_filename, file_type, final_size, current_time))
            
            uploaded_files.append({
                'id': cursor.lastrowid,
                'filename': stored_filename,
                'original_filename': original_filename,
                'file_type': file_type,
                'file_size': final_size,
                'uploaded_at': current_time,
                'was_compressed': was_compressed,
                'compression_ratio': compression_ratio
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'files': uploaded_files})
    except Exception as e:
        logging.error(f"Error uploading memo files: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/memo_files/download_all/<int:memo_id>')
def download_all_memo_files(memo_id):
    """Download all files of a memo as a ZIP archive."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Get memo info
        cursor.execute('SELECT title FROM memo WHERE id = ?', (memo_id,))
        memo_row = cursor.fetchone()
        if not memo_row:
            conn.close()
            return jsonify({'error': 'Memo not found'}), 404
        
        memo_title = memo_row[0]
        
        # Get files
        cursor.execute('''
            SELECT filename, original_filename, file_type 
            FROM memo_files 
            WHERE memo_id = ?
            ORDER BY uploaded_at ASC
        ''', (memo_id,))
        files = cursor.fetchall()
        conn.close()
        
        if not files:
            return jsonify({'error': 'No files found'}), 404
        
        # Create ZIP file in memory
        memory_file = io.BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            for filename, original_filename, file_type in files:
                file_path = os.path.join(FILEUPLOAD_DIR, filename)
                if os.path.exists(file_path):
                    # Use original filename in ZIP
                    zf.write(file_path, original_filename)
        
        memory_file.seek(0)
        
        # Create safe filename for ZIP
        safe_title = "".join(c for c in memo_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        zip_filename = f"{safe_title}_files.zip" if safe_title else f"memo_{memo_id}_files.zip"
        
        return app.response_class(
            memory_file.getvalue(),
            mimetype='application/zip',
            headers={'Content-Disposition': f'attachment; filename="{zip_filename}"'}
        )
        
    except Exception as e:
        logging.error(f"Error downloading memo files: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/memo_files/<int:file_id>', methods=['DELETE'])
def delete_memo_file(file_id):
    """Delete a memo file."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Get file info
        cursor.execute('SELECT filename FROM memo_files WHERE id = ?', (file_id,))
        file_row = cursor.fetchone()
        
        if not file_row:
            conn.close()
            return jsonify({'error': 'File not found'}), 404
        
        filename = file_row[0]
        
        # Delete from filesystem
        file_path = os.path.join(FILEUPLOAD_DIR, filename)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                logging.error(f"Error deleting file {file_path}: {str(e)}")
        
        # Delete from database
        cursor.execute('DELETE FROM memo_files WHERE id = ?', (file_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error deleting memo file: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/memo_files/<filename>')
def serve_memo_file(filename):
    """Serve memo files."""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Security check - prevent path traversal
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        return jsonify({'error': 'Invalid filename'}), 400
    
    file_path = os.path.join(FILEUPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT original_filename, file_type FROM memo_files WHERE filename = ?', (filename,))
        file_info = cursor.fetchone()
        conn.close()
        
        if file_info:
            original_filename, file_type = file_info
            
            # For image files, serve directly without download prompt
            if file_type == 'image':
                return send_file(file_path)
            
            # For non-image files, always serve as download with stored filename
            # If file was compressed, it will be served as .zip
            # If file was not compressed, it will be served with original extension
            stored_ext = os.path.splitext(filename)[1].lower()
            if stored_ext == '.zip':
                # File was compressed, serve as ZIP
                download_name = f"{os.path.splitext(original_filename)[0]}.zip"
            else:
                # File was not compressed, serve with original name
                download_name = original_filename
            
            return send_file(
                file_path, 
                as_attachment=True, 
                download_name=download_name,
                mimetype='application/octet-stream'
            )
        
        # Fallback - serve as is
        return send_file(file_path, as_attachment=True)
        
    except Exception as e:
        logging.error(f"Error serving file {filename}: {str(e)}")
        return send_file(file_path, as_attachment=True)


@app.route('/copy_project', methods=['POST'])
def copy_project():
    """Copy a project with new data."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get project name from form - ƯU TIÊN project_name
        project_name = request.form.get('project_name', '').strip()
        if not project_name:
            return jsonify({'error': 'プロジェクト名は必須です'}), 400
        
        # Connect to database first to check duplicate
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Check if project name already exists - CASE INSENSITIVE
        cursor.execute('SELECT COUNT(*) FROM projects WHERE LOWER(案件名) = LOWER(?)', (project_name,))
        existing_count = cursor.fetchone()[0]
        
        if existing_count > 0:
            conn.close()
            logging.warning(f"Duplicate project name attempt: {project_name}")
            return jsonify({'error': 'このプロジェクト名は既に存在します'}), 400
        
        # Get form data
        data = {}
        for key in request.form.keys():
            if key == 'project_name':
                # Skip project_name, chúng ta sẽ set sau
                continue
            elif key == '案件名':
                # Bỏ qua 案件名 từ form, dùng project_name thay thế
                continue
            elif key == 'タスク':
                data[key] = request.form.get(key, '')
            elif key in ['注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト']:
                data[key] = 1 if request.form.get(key) == '1' else 0
            elif key in ['開発工数（h）', '設計工数（h）']:
                value = request.form.get(key, '')
                data[key] = float(value) if value else None
            elif key == 'ページ数':
                value = request.form.get(key, '')
                data[key] = int(value) if value else None
            else:
                data[key] = request.form.get(key, '')
        
        # FORCE set 案件名 to project_name
        data['案件名'] = project_name
        
        # Ensure required fields
        if 'ステータス' not in data or not data['ステータス']:
            data['ステータス'] = '要件引継待ち'
        
        # Insert new project
        columns = []
        values = []
        placeholders = []
        
        for column in DISPLAY_COLUMNS + ['注文設計', '注文テスト', '注文FB', '注文BrSE', '並行テスト']:
            if column in data and data[column] is not None and data[column] != '':
                columns.append(f'"{column}"')
                values.append(data[column])
                placeholders.append('?')
        
        if columns:
            query = f'''
                INSERT INTO projects ({', '.join(columns)})
                VALUES ({', '.join(placeholders)})
            '''
            cursor.execute(query, values)
        
        new_project_id = cursor.lastrowid
        
        # Add history entry for copied project
        cursor.execute('''
            INSERT INTO project_history (project_id, action_type, action_details, created_at)
            VALUES (?, 'copied', 'プロジェクトコピーから作成', datetime('now', 'localtime'))
        ''', (new_project_id,))
        
        conn.commit()
        conn.close()
        
        logging.info(f"Project copied successfully with ID: {new_project_id}, name: {project_name}")
        return jsonify({'success': True, 'project_id': new_project_id})
        
    except Exception as e:
        logging.error(f"Error copying project: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/todo')
def todo():
    """Render TODO page."""
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('todo.html')

# TODO API endpoints
@app.route('/api/todos', methods=['GET'])
def get_todos():
    """Get todos for date range."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'error': 'start_date and end_date are required'}), 400
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT t.id, t.title, t.date, t.priority, t.completed, t.repeat_type, t.repeat_interval, 
                   t.repeat_unit, t.end_date, t.parent_id, t.created_at
            FROM todos t
            LEFT JOIN projects p ON t.project_id = p.id
            WHERE t.date BETWEEN ? AND ?
            AND (t.project_id IS NULL OR p.不要 != 1 OR p.不要 IS NULL)
            ORDER BY t.date, t.priority DESC, t.created_at
        ''', (start_date, end_date))
        
        todos = []
        for row in cursor.fetchall():
            todos.append({
                'id': row[0],
                'title': row[1],
                'date': row[2],
                'priority': row[3],
                'completed': bool(row[4]),
                'repeat_type': row[5],
                'repeat_interval': row[6],
                'repeat_unit': row[7],
                'end_date': row[8],
                'parent_id': row[9],
                'created_at': row[10]
            })
        
        conn.close()
        return jsonify(todos)
        
    except Exception as e:
        logging.error(f"Error fetching todos: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/todos', methods=['POST'])
def create_todo():
    """Create new todo(s) with repeat support."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    title = data.get('title', '').strip()
    date = data.get('date')
    priority = data.get('priority', 'medium')
    repeat_type = data.get('repeat_type', 'none')
    repeat_interval = data.get('repeat_interval', 1)
    repeat_unit = data.get('repeat_unit', 'days')
    end_date = data.get('end_date')
    
    if not title or not date:
        return jsonify({'error': 'Title and date are required'}), 400
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create main todo
        cursor.execute('''
            INSERT INTO todos (title, date, priority, repeat_type, repeat_interval, 
                             repeat_unit, end_date, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (title, date, priority, repeat_type, repeat_interval, repeat_unit, end_date, current_time))
        
        parent_id = cursor.lastrowid
        
        # Create repeated todos if needed
        if repeat_type != 'none' and end_date:
            current_date = datetime.strptime(date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            
            while True:
                if repeat_type == 'daily':
                    current_date += timedelta(days=repeat_interval)
                elif repeat_type == 'weekly':
                    current_date += timedelta(weeks=repeat_interval)
                elif repeat_type == 'monthly':
                    current_date += relativedelta(months=repeat_interval)
                elif repeat_type == 'custom':
                    if repeat_unit == 'days':
                        current_date += timedelta(days=repeat_interval)
                    elif repeat_unit == 'weeks':
                        current_date += timedelta(weeks=repeat_interval)
                    elif repeat_unit == 'months':
                        current_date += relativedelta(months=repeat_interval)
                
                if current_date.date() > end_date_obj.date():
                    break
                
                cursor.execute('''
                    INSERT INTO todos (title, date, priority, repeat_type, repeat_interval,
                                     repeat_unit, end_date, parent_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (title, current_date.strftime('%Y-%m-%d'), priority, repeat_type, 
                      repeat_interval, repeat_unit, end_date, parent_id, current_time))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': parent_id})
        
    except Exception as e:
        logging.error(f"Error creating todo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/todos/<int:todo_id>', methods=['PUT'])
def update_todo(todo_id):
    """Update todo."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    update_all = data.get('update_all', False)
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        if update_all:
            # Update all related todos
            cursor.execute('SELECT parent_id FROM todos WHERE id = ?', (todo_id,))
            row = cursor.fetchone()
            parent_id = row[0] if row and row[0] else todo_id
            
            update_fields = []
            values = []
            for key, value in data.items():
                if key not in ['update_all']:
                    update_fields.append(f'{key} = ?')
                    values.append(value)
            
            if update_fields:
                query = f'''
                    UPDATE todos SET {', '.join(update_fields)}
                    WHERE parent_id = ? OR id = ?
                '''
                values.extend([parent_id, parent_id])
                cursor.execute(query, values)
        else:
            # Update single todo
            update_fields = []
            values = []
            for key, value in data.items():
                if key not in ['update_all']:
                    update_fields.append(f'{key} = ?')
                    values.append(value)
            
            if update_fields:
                query = f'UPDATE todos SET {", ".join(update_fields)} WHERE id = ?'
                values.append(todo_id)
                cursor.execute(query, values)
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
        
    except Exception as e:
        logging.error(f"Error updating todo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/todos/<int:todo_id>', methods=['DELETE'])
def delete_todo(todo_id):
    """Delete todo."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    delete_all = request.args.get('delete_all', 'false').lower() == 'true'
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        if delete_all:
            # Delete all related todos
            cursor.execute('SELECT parent_id FROM todos WHERE id = ?', (todo_id,))
            row = cursor.fetchone()
            parent_id = row[0] if row and row[0] else todo_id
            
            cursor.execute('DELETE FROM todos WHERE parent_id = ? OR id = ?', (parent_id, parent_id))
        else:
            # Delete single todo
            cursor.execute('DELETE FROM todos WHERE id = ?', (todo_id,))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
        
    except Exception as e:
        logging.error(f"Error deleting todo: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
    
@app.route('/api/auto_create_todos_for_week', methods=['POST'])
def auto_create_todos_for_week():
    """Auto create TODOs from schedule data for a given week if not already created."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    week_start_str = data.get('week_start')
    
    if not week_start_str:
        return jsonify({'error': 'week_start is required'}), 400
    
    try:
        week_start = datetime.strptime(week_start_str, '%Y-%m-%d')
        week_end = week_start + timedelta(days=6)
        
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        
        # Get schedule data
        cursor.execute('''
            SELECT id, 案件名, 要件引継, 設計開始, 設計完了, 設計書送付, 開発開始, 開発完了, "PJNo.", 案件番号, 並行テスト,
                   テスト開始日, テスト完了日, FB完了予定日, SE納品, SE, PH, "開発工数（h）"
            FROM projects WHERE 不要 = 0
        ''')
        projects = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
        
        date_columns = ['要件引継', '設計開始', '設計完了', '設計書送付', '開発開始', '開発完了', 
                       'テスト開始日', 'テスト完了日', 'FB完了予定日', 'SE納品']
        
        task_mapping = {
            '要件引継': '要件引継',
            '設計開始': '設計開始', 
            '設計完了': '設計完了',
            '設計書送付': '設計書送付',
            '開発開始': '開発開始',
            '開発完了': '開発完了',
            'テスト開始日': 'テスト開始',
            'テスト完了日': 'テスト完了',
            'FB完了予定日': 'FB完了',
            'SE納品': 'SE納品'
        }
        
        priority_mapping = {
            '要件引継': 'medium',
            '設計開始': 'low',
            '設計完了': 'high',
            '設計書送付': 'high', 
            '開発開始': 'low',
            '開発完了': 'high',
            'テスト開始日': 'high',
            'テスト完了日': 'high',
            'FB完了予定日': 'high',
            'SE納品': 'high'
        }
        
        todos_created = 0
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        

        for project in projects:
            # --- 1. Tạo set milestone hợp lệ theo dữ liệu hiện tại ---
            valid_todos = set()
            parallel_test = project.get('並行テスト', 0) == 1 or project.get('並行テスト') == '1'
            parallel_todos = set()  # (title, date)
            for date_col in date_columns:
                if date_col == '設計開始':
                    continue
                date_str = project.get(date_col)
                if date_str and isinstance(date_str, str):
                    try:
                        date_obj = parse_date(date_str)
                        if week_start.date() <= date_obj.date() <= week_end.date():
                            task_name = task_mapping.get(date_col, date_col)
                            project_name = project['案件名'] or f"PJ-{project['id']}"
                            ph = project['PH'] or ''
                            ph_text = f" PH{ph}" if ph else ""
                            pjno = str(project.get('PJNo.') or project.get('案件番号') or '').strip()
                            pjno_text = f" (PJ {pjno})" if pjno else ""
                            todo_title = f"[{task_name}] {project_name}{pjno_text}{ph_text}"
                            todo_date = date_obj.strftime('%Y-%m-%d')
                            valid_todos.add((todo_title, todo_date))
                            # Nếu là テスト開始 và 並行テスト=ON, thêm SE納品(並行)
                            if task_name == 'テスト開始' and parallel_test:
                                parallel_title = f"[SE納品(並行)] {project_name}{pjno_text}{ph_text}"
                                parallel_todos.add((parallel_title, todo_date))
                    except ValueError:
                        continue

            # --- 2. Tạo set milestone TODO hiện có trong DB ---
            cursor.execute(
                "SELECT id, title, date FROM todos WHERE project_id = ? AND date >= ? AND date <= ? AND title LIKE '[%'",
                (project['id'], week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d'))
            )
            existing_todos = cursor.fetchall()
            existing_todo_set = set((row[1], row[2]) for row in existing_todos)

            # Lấy các TODO SE納品(並行) hiện có
            cursor.execute(
                "SELECT id, title, date FROM todos WHERE project_id = ? AND date >= ? AND date <= ? AND title LIKE '[SE納品(並行)]%'",
                (project['id'], week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d'))
            )
            existing_parallel_todos = cursor.fetchall()
            existing_parallel_set = set((row[1], row[2]) for row in existing_parallel_todos)

            # --- 3. Xóa TODO milestone thừa ---
            for todo_id, title, date in existing_todos:
                if (title, date) not in valid_todos:
                    # Nếu là テスト開始 và 並行テスト=ON, xóa luôn SE納品(並行) cùng ngày
                    if title.startswith('[テスト開始]') and parallel_test:
                        # Tìm và xóa SE納品(並行) cùng ngày
                        parallel_title = title.replace('[テスト開始]', '[SE納品(並行)]')
                        cursor.execute("DELETE FROM todos WHERE project_id = ? AND title = ? AND date = ?", (project['id'], parallel_title, date))
                    cursor.execute("DELETE FROM todos WHERE id = ?", (todo_id,))

            # Xóa SE納品(並行) thừa (chỉ giữ những cái đúng ngày テスト開始)
            for todo_id, title, date in existing_parallel_todos:
                if (title, date) not in parallel_todos:
                    cursor.execute("DELETE FROM todos WHERE id = ?", (todo_id,))

            # --- 4. Tạo TODO milestone còn thiếu ---
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for todo_title, todo_date in valid_todos:
                if (todo_title, todo_date) not in existing_todo_set:
                    priority = 'medium'
                    for k, v in task_mapping.items():
                        if v in todo_title:
                            priority = priority_mapping.get(k, 'medium')
                            break
                    cursor.execute(
                        "INSERT INTO todos (project_id, title, date, priority, created_at) VALUES (?, ?, ?, ?, ?)",
                        (project['id'], todo_title, todo_date, priority, current_time)
                    )
                    todos_created += 1
                    # Nếu là テスト開始 và 並行テスト=ON, tạo luôn SE納品(並行)
                    if todo_title.startswith('[テスト開始]') and parallel_test:
                        parallel_title = todo_title.replace('[テスト開始]', '[SE納品(並行)]')
                        if (parallel_title, todo_date) not in existing_parallel_set:
                            cursor.execute(
                                "INSERT INTO todos (project_id, title, date, priority, created_at) VALUES (?, ?, ?, ?, ?)",
                                (project['id'], parallel_title, todo_date, 'high', current_time)
                            )
                            todos_created += 1

            # Thêm SE納品(並行) còn thiếu (nếu có)
            for parallel_title, parallel_date in parallel_todos:
                if (parallel_title, parallel_date) not in existing_parallel_set:
                    cursor.execute(
                        "INSERT INTO todos (project_id, title, date, priority, created_at) VALUES (?, ?, ?, ?, ?)",
                        (project['id'], parallel_title, parallel_date, 'high', current_time)
                    )
                    todos_created += 1

            # --- 5. Xử lý TODO thiết kế hàng ngày ---
            # (Tương tự: tạo set valid_design_todos, xóa thừa, thêm thiếu)
            design_start_str = project.get('設計開始')
            design_end_str = project.get('設計完了')
            valid_design_todos = set()
            if design_start_str and isinstance(design_start_str, str):
                try:
                    design_start_date = parse_date(design_start_str)
                    if design_end_str and isinstance(design_end_str, str) and design_end_str.strip():
                        try:
                            design_end_date = parse_date(design_end_str) - timedelta(days=1)
                        except ValueError:
                            design_end_date = design_start_date
                    else:
                        design_end_date = design_start_date
                    current_date = design_start_date
                    while current_date <= design_end_date:
                        if week_start.date() <= current_date.date() <= week_end.date():
                            project_name = project['案件名'] or f"PJ-{project['id']}"
                            ph = project['PH'] or ''
                            ph_text = f" PH{ph}" if ph else ""
                            pjno = str(project.get('PJNo.') or project.get('案件番号') or '').strip()
                            pjno_text = f" (PJ {pjno})" if pjno else ""
                            if current_date == design_start_date:
                                todo_title = f"[設計開始] {project_name}{pjno_text}{ph_text}"
                                priority = 'medium'
                            else:
                                todo_title = f"[設計中] {project_name}{pjno_text}{ph_text}"
                                priority = 'low'
                            todo_date = current_date.strftime('%Y-%m-%d')
                            valid_design_todos.add((todo_title, todo_date, priority))
                        current_date += timedelta(days=1)
                except ValueError:
                    pass

            # Lấy TODO thiết kế hiện có
            cursor.execute(
                "SELECT id, title, date FROM todos WHERE project_id = ? AND date >= ? AND date <= ? AND (title LIKE '[設計開始]%' OR title LIKE '[設計中]%')",
                (project['id'], week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d'))
            )
            existing_design_todos = cursor.fetchall()
            existing_design_set = set((row[1], row[2]) for row in existing_design_todos)

            # Xóa TODO thiết kế thừa
            for todo_id, title, date in existing_design_todos:
                if (title, date, 'medium') not in valid_design_todos and (title, date, 'low') not in valid_design_todos:
                    cursor.execute("DELETE FROM todos WHERE id = ?", (todo_id,))

            # Thêm TODO thiết kế còn thiếu
            for todo_title, todo_date, priority in valid_design_todos:
                if (todo_title, todo_date) not in existing_design_set:
                    cursor.execute(
                        "INSERT INTO todos (project_id, title, date, priority, created_at) VALUES (?, ?, ?, ?, ?)",
                        (project['id'], todo_title, todo_date, priority, current_time)
                    )
                    todos_created += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'todos_created': todos_created,
            'todos_skipped': 0,
            'message': f'{todos_created}件のTODOを自動作成しました。' if todos_created > 0 else 'この週にはスケジュールタスクがありません。'
        })
        
    except Exception as e:
        logging.error(f"Error auto creating todos for week: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/get_current_user', methods=['GET'])
def get_current_user():
    """Get current logged in user."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    return jsonify({
        'success': True,
        'username': session['username']
    })

@app.route('/update_user_settings', methods=['POST'])
def update_user_settings():
    """Update user settings (username and password)."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        current_password = data.get('current_password', '')
        new_username = data.get('new_username', '').strip()
        new_password = data.get('new_password', '')
        
        if not current_password:
            return jsonify({'error': '現在のパスワードが必要です'}), 400
        
        if not new_username:
            return jsonify({'error': '新しいユーザー名が必要です'}), 400
        
        # Read current users
        users = read_users()
        current_username = session['username']
        
        # Verify current password
        if current_username not in users or users[current_username] != current_password:
            return jsonify({'error': '現在のパスワードが正しくありません'}), 400
        
        # Check if new username already exists (if changed)
        if new_username != current_username and new_username in users:
            return jsonify({'error': 'このユーザー名は既に存在します'}), 400
        
        # Validate new password
        if new_password and len(new_password) < 4:
            return jsonify({'error': 'パスワードは4文字以上で入力してください'}), 400
        
        # Update users dictionary
        if new_username != current_username:
            # Remove old username
            del users[current_username]
            # Add new username
            users[new_username] = new_password if new_password else current_password
            username_changed = True
        else:
            # Just update password
            users[current_username] = new_password if new_password else current_password
            username_changed = False
        
        # Write back to file
        try:
            with open('users.txt', 'w', encoding='utf-8') as f:
                for username, password in users.items():
                    f.write(f'{username}:{password}\n')
        except Exception as e:
            logging.error(f"Error writing users.txt: {str(e)}")
            return jsonify({'error': 'ファイルの書き込みに失敗しました'}), 500
        
        # Update session if username changed
        if username_changed:
            session['username'] = new_username
        
        return jsonify({
            'success': True,
            'username_changed': username_changed,
            'message': '設定が正常に更新されました'
        })
        
    except Exception as e:
        logging.error(f"Error updating user settings: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
    
    
    
#DEMO giờ dự án
@app.route('/design_hours_table')
def design_hours_table():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 案件名, 設計工数（h）, 設計開始, 設計完了, ステータス
        FROM projects
        WHERE ステータス != '不要'
    ''')
    projects = cursor.fetchall()
    conn.close()

    # Tìm tất cả ngày xuất hiện trong các dự án
    all_dates = set()
    project_rows = []
    for name, hours, start, end, status in projects:
        if not name or not hours or not start or not end:
            continue
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d')
            end_date = datetime.strptime(end, '%Y-%m-%d')
        except Exception:
            continue
        days = (end_date - start_date).days + 1
        if days <= 0:
            continue
        date_list = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days)]
        all_dates.update(date_list)
        project_rows.append({
            'project_name': name,
            'total_hours': float(hours),
            'date_list': date_list,
            'days': days
        })

    # Sắp xếp header ngày
    date_headers = sorted(list(all_dates))
    # Chuẩn bị dữ liệu từng row
    table_data = []
    for p in project_rows:
        daily = []
        # Chia đều, làm tròn 1 chữ số thập phân
        if p['days'] == 1:
            daily_hours = [round(p['total_hours'], 1)]
        else:
            avg = round(p['total_hours'] / p['days'], 1)
            daily_hours = [avg] * p['days']
            # Dồn phần dư vào ngày cuối
            total_assigned = round(avg * (p['days'] - 1), 1)
            last_day = round(p['total_hours'] - total_assigned, 1)
            daily_hours[-1] = last_day
        # Map vào từng ngày của header
        daily_map = []
        for d in date_headers:
            if d in p['date_list']:
                idx = p['date_list'].index(d)
                daily_map.append(daily_hours[idx])
            else:
                daily_map.append("")
        table_data.append({
            'project_name': p['project_name'],
            'total_hours': round(p['total_hours'], 1),
            'daily_hours': daily_map
        })

    # Tính tổng số giờ cho từng ngày
    total_hours_per_day = []
    for col_idx in range(len(date_headers)):
        s = sum(row['daily_hours'][col_idx] if isinstance(row['daily_hours'][col_idx], (int, float)) else 0 for row in table_data)
        total_hours_per_day.append(round(s, 1) if s > 0 else "")

    return render_template(
        'design_hours_table.html',
        date_headers=date_headers,
        table_data=table_data,
        total_hours_per_day=total_hours_per_day
    )

@app.template_filter('datetimeformat_jp')
def datetimeformat_jp(value):
    # value dạng yyyy-mm-dd
    import datetime
    weekdays = ['月', '火', '水', '木', '金', '土', '日']
    try:
        dt = datetime.datetime.strptime(value, '%Y-%m-%d')
        jp = dt.strftime('%Y/%m/%d')
        wd = weekdays[dt.weekday()]
        return f"{jp}({wd})"
    except Exception:
        return value

def sync_design_todo(project_id, project_name, ph, new_start, new_end, pjno=None):
    """Đồng bộ TODO [設計中] cho dự án khi update ngày."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    ph_text = f" PH{ph}" if ph else ""
    pjno_text = f" (PJ {pjno})" if pjno else ""
    todo_title = f"[設計中] {project_name}{pjno_text}{ph_text}"

    if not new_start or not new_end:
        conn.close()
        return
    start_date = datetime.strptime(new_start, '%Y-%m-%d')
    end_date = datetime.strptime(new_end, '%Y-%m-%d')
    days = (end_date - start_date).days + 1
    new_dates = set((start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days))

    # Lấy danh sách TODO [設計中] hiện có của dự án này (lọc theo project_id)
    cursor.execute('SELECT id, date FROM todos WHERE project_id = ? AND title = ?', (project_id, todo_title))
    todos = cursor.fetchall()
    old_dates = set(row[1] for row in todos)

    # Xóa TODO cũ không còn trong khoảng mới
    for todo_id, date in todos:
        if date not in new_dates:
            cursor.execute('DELETE FROM todos WHERE id = ?', (todo_id,))

    # Thêm TODO mới cho ngày chưa có
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for date in new_dates:
        if date not in old_dates:
            cursor.execute(
                "INSERT INTO todos (project_id, title, date, priority, created_at) VALUES (?, ?, ?, ?, ?)",
                (project_id, todo_title, date, 'low', current_time)
            )

    conn.commit()
    conn.close()

@app.route('/genscript')
def genscript_page():
    """Render GenScript page."""
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('genscript.html')

@app.route('/api/generate_script', methods=['POST'])
def generate_script():
    """Generate SQL script from uploaded file."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    # Get systemId from form
    system_id = request.form.get('systemId', '').strip()
    if not system_id:
        return jsonify({'success': False, 'error': 'System ID is required'}), 400
    
    if not system_id.isdigit():
        return jsonify({'success': False, 'error': 'System ID must contain numbers only'}), 400
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Invalid file format'}), 400

    try:
        # Create necessary directories
        if not os.path.exists('temp'):
            os.makedirs('temp')
        if not os.path.exists('gen_script'):
            os.makedirs('gen_script')
        
        # Save uploaded file temporarily
        temp_path = os.path.join('temp', file.filename)
        file.save(temp_path)
        
        # Generate timestamp for output file
        timestamp = datetime.now().strftime('%Y%m%d%H%M')
        base_name, ext = os.path.splitext(file.filename)
        safe_base_name = base_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
        output_filename = f'{safe_base_name}_{timestamp}.sql'
        zip_filename = f'{safe_base_name}_{timestamp}.zip'
        
        # Start generation in background thread
        def generate_background():
            global generation_progress
            try:
                generation_progress['is_generating'] = True
                generation_progress['current_sheet'] = 0
                generation_progress['total_sheets'] = 0
                generation_progress['sheet_name'] = ''
                
                # Call genscript function with systemId
                insert_statements = genscript.all_tables_in_sequence_with_progress(
                    temp_path, 
                    'table_info.txt',
                    output_filename,
                    system_id=system_id,
                    progress_callback=update_progress
                )
            
                
                # Create ZIP file
                zip_path = os.path.join('gen_script', zip_filename)
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(output_filename, output_filename)
                
                # Clean up temp files
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                if os.path.exists(output_filename):
                    os.remove(output_filename)
                    
            except Exception as e:
                logging.error(f"Error in background generation: {str(e)}")
            finally:
                generation_progress['is_generating'] = False
        
        def update_progress(current, total, sheet_name):
            global generation_progress
            generation_progress['current_sheet'] = current
            generation_progress['total_sheets'] = total
            generation_progress['sheet_name'] = sheet_name
        
        # Start background thread
        thread = threading.Thread(target=generate_background)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'filename': zip_filename,
            'message': 'Script generation started'
        })
        
    except Exception as e:
        logging.error(f"Error generating script: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generation_progress')
def get_generation_progress():
    """Get current generation progress."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    global generation_progress
    return jsonify(generation_progress)

@app.route('/api/download_script/<filename>')
def download_script(filename):
    """Download generated script file."""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Security check
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        return jsonify({'error': 'Invalid filename'}), 400
    
    file_path = os.path.join('gen_script', filename)
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        return send_file(file_path, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete_all_scripts', methods=['DELETE'])
def delete_all_scripts():
    """Delete all generated script files."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        gen_script_dir = 'gen_script'
        if os.path.exists(gen_script_dir):
            for filename in os.listdir(gen_script_dir):
                file_path = os.path.join(gen_script_dir, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
        
        return jsonify({'success': True, 'message': 'All scripts deleted successfully'})
        
    except Exception as e:
        logging.error(f"Error deleting scripts: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/script_files')
def get_script_files():
    """Get list of generated script files."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        gen_script_dir = 'gen_script'
        files = []
        
        if os.path.exists(gen_script_dir):
            for filename in os.listdir(gen_script_dir):
                file_path = os.path.join(gen_script_dir, filename)
                if os.path.isfile(file_path) and filename.endswith('.zip'):
                    files.append(filename)
        
        files.sort(reverse=True)  # Newest first
        return jsonify({'files': files})
        
    except Exception as e:
        logging.error(f"Error getting script files: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/get_configuration')
def get_configuration():
    """Get current configuration files content."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        table_info_content = ''
        gendoc_config_content = ''
        
        # Read table_info.txt
        if os.path.exists('table_info.txt'):
            with open('table_info.txt', 'r', encoding='utf-8') as f:
                table_info_content = f.read()
        
        # Read gendoc_config.txt
        if os.path.exists('gendoc_config.txt'):
            with open('gendoc_config.txt', 'r', encoding='utf-8') as f:
                gendoc_config_content = f.read()
        
        # Also parse for detailed view
        config_data = {
            'table_info': {},
            'mapping_values': {},
            'koumoku_mapping': {},
            'koumoku_mapping_re': {},
            'stop_values': [],
            'excluded_sheets': [],
            'merged_cells': {},
            'skip_cells': [],
            'row_processor': {},
            'logic_processor': {}
        }
        
        if table_info_content.strip():
            try:
                config_data['table_info'] = json.loads(table_info_content)
            except json.JSONDecodeError:
                pass
        
        if gendoc_config_content.strip():
            try:
                gendoc_config = json.loads(gendoc_config_content)
                config_data['mapping_values'] = gendoc_config.get('MAPPING_VALUE_DICT', {})
                config_data['koumoku_mapping'] = gendoc_config.get('KOUMOKU_TYPE_MAPPING', {})
                config_data['koumoku_mapping_re'] = gendoc_config.get('KOUMOKU_TYPE_MAPPING_RE', {})
                config_data['stop_values'] = gendoc_config.get('STOP_VALUES', [])
                config_data['excluded_sheets'] = gendoc_config.get('EXCLUDED_SHEETNAMES', [])
                config_data['merged_cells'] = gendoc_config.get('MERGED_CELL_RANGES', {})
                config_data['skip_cells'] = gendoc_config.get('SKIP_CELL_VALUES', [])
                config_data['row_processor'] = gendoc_config.get('ROW_PROCESSOR_CONFIG', {})
                config_data['logic_processor'] = gendoc_config.get('LOGIC_PROCESSOR_CONFIG', {})
            except json.JSONDecodeError:
                pass
        
        return jsonify({
            'success': True,
            'table_info': table_info_content,
            'gendoc_config': gendoc_config_content,
            'detailed_config': config_data
        })
        
    except Exception as e:
        logging.error(f"Error getting configuration: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/update_configuration', methods=['POST'])
def update_configuration():
    """Update configuration files."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        table_info = data.get('table_info', '')
        gendoc_config = data.get('gendoc_config', '')
        
        # Validate JSON format for gendoc_config
        if gendoc_config.strip():
            try:
                json.loads(gendoc_config)
            except json.JSONDecodeError as e:
                return jsonify({'success': False, 'error': f'Invalid JSON in gendoc_config: {str(e)}'}), 400
        
        # Validate JSON format for table_info
        if table_info.strip():
            try:
                json.loads(table_info)
            except json.JSONDecodeError as e:
                return jsonify({'success': False, 'error': f'Invalid JSON in table_info: {str(e)}'}), 400
        
        # Write table_info.txt
        if table_info.strip():
            with open('table_info.txt', 'w', encoding='utf-8') as f:
                f.write(table_info)
        
        # Write gendoc_config.txt
        if gendoc_config.strip():
            with open('gendoc_config.txt', 'w', encoding='utf-8') as f:
                f.write(gendoc_config)
        
        return jsonify({'success': True, 'message': 'Configuration updated successfully'})
        
    except Exception as e:
        logging.error(f"Error updating configuration: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Thêm các routes mới để xử lý cấu hình theo nhóm

@app.route('/api/get_detailed_configuration')
def get_detailed_configuration():
    """Get configuration broken down by sections for detailed editing."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        config_data = {
            'table_info': {},
            'mapping_values': {},
            'koumoku_mapping': {},
            'koumoku_mapping_re': {},
            'stop_values': [],
            'excluded_sheets': [],
            'merged_cells': {},
            'skip_cells': [],
            'row_processor': {},
            'logic_processor': {}
        }
        
        # Read table_info.txt
        if os.path.exists('table_info.txt'):
            try:
                with open('table_info.txt', 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    if content:
                        config_data['table_info'] = json.loads(content)
            except (json.JSONDecodeError, Exception) as e:
                logging.error(f"Error parsing table_info.txt: {str(e)}")
        
        # Read gendoc_config.txt
        if os.path.exists('gendoc_config.txt'):
            try:
                with open('gendoc_config.txt', 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    if content:
                        gendoc_config = json.loads(content)
                        config_data['mapping_values'] = gendoc_config.get('MAPPING_VALUE_DICT', {})
                        config_data['koumoku_mapping'] = gendoc_config.get('KOUMOKU_TYPE_MAPPING', {})
                        config_data['koumoku_mapping_re'] = gendoc_config.get('KOUMOKU_TYPE_MAPPING_RE', {})
                        config_data['stop_values'] = gendoc_config.get('STOP_VALUES', [])
                        config_data['excluded_sheets'] = gendoc_config.get('EXCLUDED_SHEETNAMES', [])
                        config_data['merged_cells'] = gendoc_config.get('MERGED_CELL_RANGES', {})
                        config_data['skip_cells'] = gendoc_config.get('SKIP_CELL_VALUES', [])
                        config_data['row_processor'] = gendoc_config.get('ROW_PROCESSOR_CONFIG', {})
                        config_data['logic_processor'] = gendoc_config.get('LOGIC_PROCESSOR_CONFIG', {})
            except (json.JSONDecodeError, Exception) as e:
                logging.error(f"Error parsing gendoc_config.txt: {str(e)}")
        
        return jsonify({
            'success': True,
            'config': config_data
        })
        
    except Exception as e:
        logging.error(f"Error getting detailed configuration: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/update_detailed_configuration', methods=['POST'])
def update_detailed_configuration():
    """Update configuration from detailed form data."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        
        # Extract configuration sections
        table_info = data.get('table_info', {})
        mapping_values = data.get('mapping_values', {})
        koumoku_mapping = data.get('koumoku_mapping', {})
        koumoku_mapping_re = data.get('koumoku_mapping_re', {})
        stop_values = data.get('stop_values', [])
        excluded_sheets = data.get('excluded_sheets', [])
        merged_cells = data.get('merged_cells', {})
        skip_cells = data.get('skip_cells', [])
        row_processor = data.get('row_processor', {})
        logic_processor = data.get('logic_processor', {})
        
        # Validate table_info JSON
        if table_info:
            try:
                # Validate each table configuration
                for table_name, table_config in table_info.items():
                    if not isinstance(table_config, list):
                        return jsonify({'success': False, 'error': f'Table {table_name} configuration must be an array'}), 400
                    # Validate each column configuration
                    for col_config in table_config:
                        if not isinstance(col_config, dict):
                            return jsonify({'success': False, 'error': f'Column configuration in table {table_name} must be an object'}), 400
            except Exception as e:
                return jsonify({'success': False, 'error': f'Invalid table_info format: {str(e)}'}), 400
        
        # Validate complex objects (row_processor, logic_processor)
        for config_name, config_data in [('row_processor', row_processor), ('logic_processor', logic_processor)]:
            if config_data:
                try:
                    # Ensure it's valid JSON-serializable
                    json.dumps(config_data)
                except Exception as e:
                    return jsonify({'success': False, 'error': f'Invalid {config_name} format: {str(e)}'}), 400
        
        # Build gendoc_config
        gendoc_config = {
            'MAPPING_VALUE_DICT': mapping_values,
            'KOUMOKU_TYPE_MAPPING': koumoku_mapping,
            'KOUMOKU_TYPE_MAPPING_RE': koumoku_mapping_re,
            'STOP_VALUES': stop_values,
            'EXCLUDED_SHEETNAMES': excluded_sheets,
            'MERGED_CELL_RANGES': merged_cells,
            'SKIP_CELL_VALUES': skip_cells,
            'ROW_PROCESSOR_CONFIG': row_processor,
            'LOGIC_PROCESSOR_CONFIG': logic_processor
        }
        
        # Write files
        if table_info:
            with open('table_info.txt', 'w', encoding='utf-8') as f:
                json.dump(table_info, f, ensure_ascii=False, indent=2)
        
        with open('gendoc_config.txt', 'w', encoding='utf-8') as f:
            json.dump(gendoc_config, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': 'Configuration updated successfully'})
        
    except Exception as e:
        logging.error(f"Error updating detailed configuration: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/add_config_item', methods=['POST'])
def add_config_item():
    """Add a new configuration item to a specific section."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        section = data.get('section')
        key = data.get('key', '').strip()
        value = data.get('value', '')
        
        if not section or not key:
            return jsonify({'success': False, 'error': 'Section and key are required'}), 400
        
        # Read current configuration
        config_data = {}
        if os.path.exists('gendoc_config.txt'):
            with open('gendoc_config.txt', 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    config_data = json.loads(content)
        
        # Add new item based on section type
        if section in ['mapping_values', 'koumoku_mapping', 'koumoku_mapping_re', 'merged_cells']:
            section_map = {
                'mapping_values': 'MAPPING_VALUE_DICT',
                'koumoku_mapping': 'KOUMOKU_TYPE_MAPPING',
                'koumoku_mapping_re': 'KOUMOKU_TYPE_MAPPING_RE',
                'merged_cells': 'MERGED_CELL_RANGES'
            }
            config_key = section_map[section]
            if config_key not in config_data:
                config_data[config_key] = {}
            config_data[config_key][key] = value
            
        elif section in ['stop_values', 'excluded_sheets', 'skip_cells']:
            section_map = {
                'stop_values': 'STOP_VALUES',
                'excluded_sheets': 'EXCLUDED_SHEETNAMES',
                'skip_cells': 'SKIP_CELL_VALUES'
            }
            config_key = section_map[section]
            if config_key not in config_data:
                config_data[config_key] = []
            if key not in config_data[config_key]:
                config_data[config_key].append(key)
                
        elif section in ['row_processor', 'logic_processor']:
            section_map = {
                'row_processor': 'ROW_PROCESSOR_CONFIG',
                'logic_processor': 'LOGIC_PROCESSOR_CONFIG'
            }
            config_key = section_map[section]
            if config_key not in config_data:
                config_data[config_key] = {}
            # Try to parse value as JSON for complex objects
            try:
                parsed_value = json.loads(value) if isinstance(value, str) else value
                config_data[config_key][key] = parsed_value
            except json.JSONDecodeError:
                config_data[config_key][key] = value
        
        # Write back to file
        with open('gendoc_config.txt', 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': f'Added {key} to {section}'})
        
    except Exception as e:
        logging.error(f"Error adding config item: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/remove_config_item', methods=['POST'])
def remove_config_item():
    """Remove a configuration item from a specific section."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        section = data.get('section')
        key = data.get('key', '').strip()
        
        if not section or not key:
            return jsonify({'success': False, 'error': 'Section and key are required'}), 400
        
        # Read current configuration
        config_data = {}
        if os.path.exists('gendoc_config.txt'):
            with open('gendoc_config.txt', 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    config_data = json.loads(content)
        
        # Remove item based on section type
        removed = False
        if section in ['mapping_values', 'koumoku_mapping', 'koumoku_mapping_re', 'merged_cells']:
            section_map = {
                'mapping_values': 'MAPPING_VALUE_DICT',
                'koumoku_mapping': 'KOUMOKU_TYPE_MAPPING',
                'koumoku_mapping_re': 'KOUMOKU_TYPE_MAPPING_RE',
                'merged_cells': 'MERGED_CELL_RANGES'
            }
            config_key = section_map[section]
            if config_key in config_data and key in config_data[config_key]:
                del config_data[config_key][key]
                removed = True
                
        elif section in ['stop_values', 'excluded_sheets', 'skip_cells']:
            section_map = {
                'stop_values': 'STOP_VALUES',
                'excluded_sheets': 'EXCLUDED_SHEETNAMES',
                'skip_cells': 'SKIP_CELL_VALUES'
            }
            config_key = section_map[section]
            if config_key in config_data and key in config_data[config_key]:
                config_data[config_key].remove(key)
                removed = True
                
        elif section in ['row_processor', 'logic_processor']:
            section_map = {
                'row_processor': 'ROW_PROCESSOR_CONFIG',
                'logic_processor': 'LOGIC_PROCESSOR_CONFIG'
            }
            config_key = section_map[section]
            if config_key in config_data and key in config_data[config_key]:
                del config_data[config_key][key]
                removed = True
        
        if not removed:
            return jsonify({'success': False, 'error': f'Item {key} not found in {section}'}), 404
        
        # Write back to file
        with open('gendoc_config.txt', 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': f'Removed {key} from {section}'})
        
    except Exception as e:
        logging.error(f"Error removing config item: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/add_table_config', methods=['POST'])
def add_table_config():
    """Add a new table configuration."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        table_name = data.get('table_name', '').strip()
        table_config = data.get('table_config', [])
        
        if not table_name:
            return jsonify({'success': False, 'error': 'Table name is required'}), 400
        
        # Validate table_config format
        if not isinstance(table_config, list):
            return jsonify({'success': False, 'error': 'Table configuration must be an array'}), 400
        
        # Read current table_info
        table_info = {}
        if os.path.exists('table_info.txt'):
            with open('table_info.txt', 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    table_info = json.loads(content)
        
        # Add new table
        table_info[table_name] = table_config
        
        # Write back to file
        with open('table_info.txt', 'w', encoding='utf-8') as f:
            json.dump(table_info, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': f'Added table {table_name}'})
        
    except Exception as e:
        logging.error(f"Error adding table config: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/remove_table_config', methods=['POST'])
def remove_table_config():
    """Remove a table configuration."""
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        table_name = data.get('table_name', '').strip()
        
        if not table_name:
            return jsonify({'success': False, 'error': 'Table name is required'}), 400
        
        # Read current table_info
        table_info = {}
        if os.path.exists('table_info.txt'):
            with open('table_info.txt', 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    table_info = json.loads(content)
        
        # Remove table
        if table_name not in table_info:
            return jsonify({'success': False, 'error': f'Table {table_name} not found'}), 404
        
        del table_info[table_name]
        
        # Write back to file
        with open('table_info.txt', 'w', encoding='utf-8') as f:
            json.dump(table_info, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': f'Removed table {table_name}'})
        
    except Exception as e:
        logging.error(f"Error removing table config: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get_placeholders_content', methods=['GET'])
def get_placeholders_content():
    """Get the content of placeholders.txt file."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        placeholders_file = 'placeholders.txt'
        if os.path.exists(placeholders_file):
            with open(placeholders_file, 'r', encoding='utf-8') as f:
                content = f.read()
            return jsonify({'content': content})
        else:
            return jsonify({'content': ''})
    except Exception as e:
        logging.error(f"Error reading placeholders.txt: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/save_placeholders_content', methods=['POST'])
def save_placeholders_content():
    """Save content to placeholders.txt file."""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.json
        content = data.get('content', '')
        
        placeholders_file = 'placeholders.txt'
        with open(placeholders_file, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logging.info(f"Successfully saved placeholders.txt")
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error saving placeholders.txt: {str(e)}")
        return jsonify({'error': str(e)}), 500

def compare_excel_files_background(task_id, old_file_path, new_file_path):
    """Background function to compare Excel files"""
    try:
        # Create temp directory for comparison results
        temp_dir = f'temp_compare_{task_id}'
        os.makedirs(temp_dir, exist_ok=True)
        
        # Create output file paths
        file_old_highlighted = os.path.join(temp_dir, 'old_highlighted.xlsx')
        file_new_highlighted = os.path.join(temp_dir, 'new_highlighted.xlsx')
        output_file = os.path.join(temp_dir, 'comparison_result.txt')
        
        # Initialize task status
        comparison_tasks[task_id].update({
            'current_sheet': 0,
            'total_sheets': 0,
            'current_sheet_name': '',
            'status': 'Loading files...'
        })
        
        # Create yellow fill for highlighting
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Copy files to preserve format
        shutil.copy2(old_file_path, file_old_highlighted)
        shutil.copy2(new_file_path, file_new_highlighted)
        
        # Load workbooks
        comparison_tasks[task_id]['status'] = 'Loading workbooks...'
        wb_old = load_workbook(old_file_path, data_only=True)
        wb_new = load_workbook(new_file_path, data_only=True)
        wb_old_highlighted = load_workbook(file_old_highlighted)
        wb_new_highlighted = load_workbook(file_new_highlighted)
        
        sheets_with_differences = []
        
        with open(output_file, 'w', encoding='utf-8') as f:
            sheets_old = wb_old.sheetnames
            sheets_new = wb_new.sheetnames
            
            # Update total sheets count
            common_sheets = set(sheets_old) & set(sheets_new)
            comparison_tasks[task_id]['total_sheets'] = len(common_sheets)
            
            # Write initial comparison info
            if len(sheets_old) != len(sheets_new):
                f.write("Số lượng sheet không giống nhau!\n")
                f.write(f"File old có {len(sheets_old)} sheets: {', '.join(sheets_old)}\n")
                f.write(f"File new có {len(sheets_new)} sheets: {', '.join(sheets_new)}\n\n")
            else:
                f.write("Số lượng sheet giống nhau. Kết quả so sánh:\n\n")
                
            only_in_old = set(sheets_old) - set(sheets_new)
            only_in_new = set(sheets_new) - set(sheets_old)
            if only_in_old:
                f.write(f"Sheet chỉ có trong file old: {', '.join(only_in_old)}\n")
            if only_in_new:
                f.write(f"Sheet chỉ có trong file new: {', '.join(only_in_new)}\n")
            if only_in_old or only_in_new:
                f.write("\n")
            
            # Process each common sheet
            for i, sheet_name in enumerate(common_sheets):
                comparison_tasks[task_id].update({
                    'current_sheet': i + 1,
                    'current_sheet_name': sheet_name,
                    'status': f'Processing sheet: {sheet_name}'
                })
                
                try:
                    ws_old = wb_old[sheet_name]
                    ws_new = wb_new[sheet_name]
                    ws_old_highlighted = wb_old_highlighted[sheet_name]
                    ws_new_highlighted = wb_new_highlighted[sheet_name]
                    
                    max_row = max(ws_old.max_row, ws_new.max_row)
                    max_col = max(ws_old.max_column, ws_new.max_column)
                    different_cells = 0
                    different_positions = []
                    has_differences = False
                    
                    # Compare cells
                    for row in range(1, max_row + 1):
                        for col in range(1, max_col + 1):
                            value_old = ws_old.cell(row=row, column=col).value
                            value_new = ws_new.cell(row=row, column=col).value
                            
                            clean_value_old = str(value_old).strip() if value_old is not None else ''
                            clean_value_new = str(value_new).strip() if value_new is not None else ''
                            
                            if clean_value_old != clean_value_new:
                                different_cells += 1
                                has_differences = True
                                
                                # Highlight different cells
                                try:
                                    cell_old = ws_old_highlighted.cell(row=row, column=col)
                                    cell_new = ws_new_highlighted.cell(row=row, column=col)
                                    
                                    # Handle merged cells
                                    merged_old = False
                                    for merged_range in ws_old_highlighted.merged_cells.ranges:
                                        if cell_old.coordinate in merged_range:
                                            for merged_cell in merged_range.cells:
                                                ws_old_highlighted.cell(row=merged_cell[0], column=merged_cell[1]).fill = yellow_fill
                                            merged_old = True
                                            break
                                    if not merged_old:
                                        cell_old.fill = yellow_fill
                                    
                                    merged_new = False
                                    for merged_range in ws_new_highlighted.merged_cells.ranges:
                                        if cell_new.coordinate in merged_range:
                                            for merged_cell in merged_range.cells:
                                                ws_new_highlighted.cell(row=merged_cell[0], column=merged_cell[1]).fill = yellow_fill
                                            merged_new = True
                                            break
                                    if not merged_new:
                                        cell_new.fill = yellow_fill
                                        
                                except Exception as e:
                                    pass
                                
                                # Record difference
                                if len(different_positions) < 20:
                                    col_letter = ws_old.cell(row=row, column=col).column_letter
                                    different_positions.append({
                                        'position': f"{col_letter}{row}",
                                        'value_old': clean_value_old if clean_value_old != '' else "NULL",
                                        'value_new': clean_value_new if clean_value_new != '' else "NULL"
                                    })
                    
                    # Collect all values for comparison
                    values_old = set()
                    values_new = set()
                    non_empty_rows_old = 0
                    non_empty_rows_new = 0
                    
                    for row in range(1, ws_old.max_row + 1):
                        has_data = False
                        for col in range(1, ws_old.max_column + 1):
                            value = ws_old.cell(row=row, column=col).value
                            if value is not None and str(value).strip() != '':
                                clean_value = str(value).strip()
                                values_old.add(clean_value)
                                has_data = True
                        if has_data:
                            non_empty_rows_old += 1
                    
                    for row in range(1, ws_new.max_row + 1):
                        has_data = False
                        for col in range(1, ws_new.max_column + 1):
                            value = ws_new.cell(row=row, column=col).value
                            if value is not None and str(value).strip() != '':
                                clean_value = str(value).strip()
                                values_new.add(clean_value)
                                has_data = True
                        if has_data:
                            non_empty_rows_new += 1
                    
                    # Write results
                    if has_differences or values_old != values_new:
                        sheets_with_differences.append(sheet_name)
                        f.write(f"Sheet: {sheet_name}\n")
                        f.write(f"Số dòng có dữ liệu trong file old: {non_empty_rows_old}\n")
                        f.write(f"Số dòng có dữ liệu trong file new: {non_empty_rows_new}\n")
                        f.write(f"Tổng số giá trị khác rỗng trong file old: {len(values_old)}\n")
                        f.write(f"Tổng số giá trị khác rỗng trong file new: {len(values_new)}\n")
                        f.write(f"Số ô có nội dung khác nhau: {different_cells}\n")
                        
                        if different_positions:
                            f.write(f"Ví dụ {min(len(different_positions), 20)} vị trí khác nhau đầu tiên:\n")
                            for pos in different_positions:
                                f.write(f"  {pos['position']}: old='{pos['value_old']}' vs new='{pos['value_new']}'\n")
                        
                        only_in_old = values_old - values_new
                        only_in_new = values_new - values_old
                        
                        if only_in_old:
                            f.write(f"Giá trị chỉ có trong file old ({len(only_in_old)} giá trị):\n")
                            for value in sorted(list(only_in_old)[:20]):
                                f.write(f"  {value}\n")
                            if len(only_in_old) > 20:
                                f.write(f"  ... và {len(only_in_old) - 20} giá trị khác\n")
                        
                        if only_in_new:
                            f.write(f"Giá trị chỉ có trong file new ({len(only_in_new)} giá trị):\n")
                            for value in sorted(list(only_in_new)[:20]):
                                f.write(f"  {value}\n")
                            if len(only_in_new) > 20:
                                f.write(f"  ... và {len(only_in_new) - 20} giá trị khác\n")
                        
                        f.write("\n" + "="*50 + "\n")
                    else:
                        f.write(f"Sheet: {sheet_name}\n")
                        f.write("Nội dung giống hệt nhau (bỏ qua dòng trắng và cấu trúc).\n")
                        f.write(f"Số ô có nội dung khác nhau: {different_cells}\n")
                        f.write("\n" + "="*50 + "\n")
                        
                except Exception as e:
                    f.write(f"Sheet: {sheet_name}\n")
                    f.write(f"Lỗi khi xử lý sheet: {e}\n")
                    f.write("\n" + "="*50 + "\n")
        
        # Remove sheets without differences from highlighted files
        for sheet_name in list(wb_old_highlighted.sheetnames):
            if sheet_name not in sheets_with_differences:
                wb_old_highlighted.remove(wb_old_highlighted[sheet_name])
        
        for sheet_name in list(wb_new_highlighted.sheetnames):
            if sheet_name not in sheets_with_differences:
                wb_new_highlighted.remove(wb_new_highlighted[sheet_name])
        
        # Close original workbooks
        wb_old.close()
        wb_new.close()
        
        # Save highlighted files
        result_files = {}
        if sheets_with_differences:
            wb_old_highlighted.save(file_old_highlighted)
            wb_new_highlighted.save(file_new_highlighted)
            result_files['old_highlighted'] = f'old_highlighted_{task_id}.xlsx'
            result_files['new_highlighted'] = f'new_highlighted_{task_id}.xlsx'
            
            # Move files to a permanent location with task_id
            shutil.move(file_old_highlighted, f'temp_compare/old_highlighted_{task_id}.xlsx')
            shutil.move(file_new_highlighted, f'temp_compare/new_highlighted_{task_id}.xlsx')
        
        wb_old_highlighted.close()
        wb_new_highlighted.close()
        
        # Move result file
        result_files['comparison_result'] = f'comparison_result_{task_id}.txt'
        os.makedirs('temp_compare', exist_ok=True)
        shutil.move(output_file, f'temp_compare/comparison_result_{task_id}.txt')
        
        # Read result text for preview
        with open(f'temp_compare/comparison_result_{task_id}.txt', 'r', encoding='utf-8') as f:
            result_text = f.read()
        
        # Update task status
        comparison_tasks[task_id].update({
            'completed': True,
            'status': 'Completed',
            'result_text': result_text[:2000] + ('...' if len(result_text) > 2000 else ''),  # Truncate for display
            'files': result_files,
            'sheets_with_differences': len(sheets_with_differences)
        })
        
        # Clean up temp directory
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
            
    except Exception as e:
        comparison_tasks[task_id].update({
            'completed': True,
            'status': 'Error',
            'error': str(e)
        })
        logging.error(f"Error in compare_excel_files_background: {str(e)}")

@app.route('/api/compare_excel', methods=['POST'])
def compare_excel():
    """Start Excel comparison process"""
    try:
        if 'old_file' not in request.files or 'new_file' not in request.files:
            return jsonify({'error': 'Both files are required'}), 400
        
        old_file = request.files['old_file']
        new_file = request.files['new_file']
        
        if old_file.filename == '' or new_file.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400
        
        # Generate unique task ID
        task_id = str(uuid.uuid4())
        
        # Save uploaded files temporarily
        temp_dir = f'temp_upload_{task_id}'
        os.makedirs(temp_dir, exist_ok=True)
        
        old_file_path = os.path.join(temp_dir, 'old.xlsx')
        new_file_path = os.path.join(temp_dir, 'new.xlsx')
        
        old_file.save(old_file_path)
        new_file.save(new_file_path)
        
        # Initialize task tracking
        comparison_tasks[task_id] = {
            'completed': False,
            'current_sheet': 0,
            'total_sheets': 0,
            'current_sheet_name': '',
            'status': 'Initializing...',
            'temp_dir': temp_dir
        }
        
        # Start background comparison
        thread = threading.Thread(target=compare_excel_files_background, args=(task_id, old_file_path, new_file_path))
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'task_id': task_id})
        
    except Exception as e:
        logging.error(f"Error in compare_excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/compare_progress/<task_id>')
def get_compare_progress(task_id):
    """Get comparison progress"""
    if task_id not in comparison_tasks:
        return jsonify({'error': 'Task not found'}), 404
    
    task_data = comparison_tasks[task_id].copy()
    
    # Clean up completed tasks after some time
    if task_data.get('completed', False):
        # Clean up temp upload directory
        temp_dir = task_data.get('temp_dir')
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
    
    return jsonify(task_data)

@app.route('/api/download_compare_file/<filename>')
def download_compare_file(filename):
    """Download comparison result files"""
    try:
        file_path = os.path.join('temp_compare', filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        logging.error(f"Error downloading file: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear_compare_files', methods=['POST'])
def clear_compare_files():
    """Clear all files in temp_compare directory"""
    try:
        temp_compare_dir = 'temp_compare'
        if os.path.exists(temp_compare_dir):
            # Remove all files in the directory
            for filename in os.listdir(temp_compare_dir):
                file_path = os.path.join(temp_compare_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    logging.warning(f"Could not delete {file_path}: {str(e)}")
        
        # Clear comparison tasks
        global comparison_tasks
        comparison_tasks.clear()
        
        return jsonify({'success': True, 'message': 'All comparison files cleared'})
        
    except Exception as e:
        logging.error(f"Error clearing compare files: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
if __name__ == '__main__':
    app.run(debug=True)